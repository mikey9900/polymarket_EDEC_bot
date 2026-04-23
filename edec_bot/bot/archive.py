"""Daily archive export: 24h Excel + compressed recent trades + optional Dropbox sync."""

from __future__ import annotations

import argparse
import base64
import csv
import gzip
import json
import logging
import os
import shutil
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib import request
from urllib import error as urlerror
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bot.export import _auto_width, _freeze, _style_header

logger = logging.getLogger(__name__)

_OPTIONAL_LATEST_KEYS = {"latest_signals_csv_gz"}

_SESSION_TRADE_COMPACT_NAMES = {
    "trade_id": "id", "timestamp": "ts", "run_id": "rid", "app_version": "av",
    "strategy_version": "sv", "config_hash": "ch", "mode": "md", "dry_run": "dr",
    "order_size_usd": "os", "paper_capital_total": "cap", "market_slug": "mkt",
    "window_id": "wid", "coin": "c", "strategy_type": "st", "signal_context": "ctx",
    "signal_overlap_count": "ov", "side": "sd", "entry_price": "ep", "entry_bid": "eb",
    "entry_ask": "ea", "entry_spread": "es", "target_price": "tp", "shares": "sh",
    "shares_requested": "srq", "shares_filled": "sfl", "blocked_min_5_shares": "b5",
    "cost": "cs", "fee_total": "fee", "status": "status", "exit_price": "xp",
    "pnl": "pnl", "exit_reason": "er", "exit_timestamp": "xt", "time_remaining_s": "tx",
    "bid_at_exit": "xb", "ask_at_exit": "xa", "exit_spread": "xs", "market_start_time": "ms",
    "market_end_time": "me", "entry_depth_side_usd": "eds", "opposite_depth_usd": "ods",
    "depth_ratio": "drt", "max_bid_seen": "maxb", "min_bid_seen": "minb",
    "time_to_max_bid_s": "ttmax", "time_to_min_bid_s": "ttmin", "first_profit_time_s": "tfp",
    "signal_score": "sg", "score_velocity": "sgv", "score_entry": "sge",
    "score_depth": "sgd", "score_spread": "sgs", "score_time": "sgt",
    "score_balance": "sgb", "target_delta": "td", "hard_stop_delta": "hsd",
    "scalp_hit": "sc", "high_confidence_hit": "hc", "filter_passed": "fp",
    "filter_failed": "ff", "decision_reason": "why", "coin_velocity_30s": "v30",
    "research_cluster_id": "rcid", "research_cluster_n": "rcn",
    "research_cluster_win_pct": "rcw", "research_cluster_avg_pnl": "rcp",
    "research_policy_action": "rpa",
    "coin_velocity_60s": "v60", "up_depth_usd": "du", "down_depth_usd": "dd",
    "decision_time_remaining_s": "te", "mfe": "mfe", "mae": "mae",
    "peak_net_pnl": "pnp", "trough_net_pnl": "tnp", "stall_exit_triggered": "sx",
    "hold_to_resolution": "hr", "loss_cut_threshold_pct": "lct",
    "loss_pct_at_exit": "lpx", "favorable_excursion": "fex", "ever_profitable": "evp",
    "source_prices_json": "spj", "source_ages_json": "saj", "source_dispersion_pct": "sdp",
    "source_staleness_max_s": "ssx", "source_staleness_avg_s": "ssa",
    "resolution_winner": "rw", "resolution_side_match": "rsm",
    "resolution_value": "rv", "resolution_pnl": "rpn", "would_have_won": "whw",
    "would_have_beaten_exit": "wbe", "missed_upside_after_exit": "mux",
    "time_to_resolution_s": "ttr", "recovered_after_exit": "rae",
}

_SESSION_SIGNAL_COMPACT_NAMES = {
    "decision_id": "id", "timestamp": "ts", "run_id": "rid", "app_version": "av",
    "strategy_version": "sv", "config_hash": "ch", "mode": "md", "dry_run": "dr",
    "order_size_usd": "os", "paper_capital_total": "cap", "market_slug": "mkt",
    "window_id": "wid", "coin": "c", "strategy_type": "st", "action": "act",
    "suppressed_reason": "sup", "reason": "why", "signal_context": "ctx",
    "signal_overlap_count": "ov", "entry_price": "ep", "target_price": "tp",
    "expected_profit_per_share": "eps", "signal_score": "sg", "score_velocity": "sgv",
    "score_entry": "sge", "score_depth": "sgd", "score_spread": "sgs",
    "score_time": "sgt", "score_balance": "sgb", "time_remaining_s": "te",
    "coin_velocity_30s": "v30", "coin_velocity_60s": "v60", "entry_bid": "eb",
    "entry_ask": "ea", "entry_spread": "es", "entry_depth_side_usd": "eds",
    "opposite_depth_usd": "ods", "depth_ratio": "drt", "resignal_cooldown_s": "rcd",
    "min_price_improvement": "mpi", "last_signal_age_s": "lsa",
    "research_cluster_id": "rcid", "research_cluster_n": "rcn",
    "research_cluster_win_pct": "rcw", "research_cluster_avg_pnl": "rcp",
    "research_policy_action": "rpa",
    "filter_passed": "fp", "filter_failed": "ff", "source_prices_json": "spj",
    "source_ages_json": "saj", "source_dispersion_pct": "sdp",
    "source_staleness_max_s": "ssx", "source_staleness_avg_s": "ssa",
}

_SESSION_TITLE_FILL = PatternFill(start_color="101D42", end_color="101D42", fill_type="solid")
_SESSION_ACCENT_FILL = PatternFill(start_color="162B64", end_color="162B64", fill_type="solid")
_SESSION_PANEL_FILL = PatternFill(start_color="EEF5FF", end_color="EEF5FF", fill_type="solid")
_SESSION_SUBHEADER_FILL = PatternFill(start_color="DCEBFF", end_color="DCEBFF", fill_type="solid")
_SESSION_POSITIVE_FILL = PatternFill(start_color="D9F7E8", end_color="D9F7E8", fill_type="solid")
_SESSION_NEGATIVE_FILL = PatternFill(start_color="FDE2E2", end_color="FDE2E2", fill_type="solid")
_SESSION_WARNING_FILL = PatternFill(start_color="FFF0D6", end_color="FFF0D6", fill_type="solid")
_SESSION_NEUTRAL_FILL = PatternFill(start_color="F4F7FB", end_color="F4F7FB", fill_type="solid")
_SESSION_BORDER = Border(
    left=Side(style="thin", color="C7D5F0"),
    right=Side(style="thin", color="C7D5F0"),
    top=Side(style="thin", color="C7D5F0"),
    bottom=Side(style="thin", color="C7D5F0"),
)


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _build_dropbox_auth(
    dropbox_token: str | None = None,
    dropbox_refresh_token: str | None = None,
    dropbox_app_key: str | None = None,
    dropbox_app_secret: str | None = None,
) -> dict[str, Any] | None:
    auth = {
        "access_token": (dropbox_token or "").strip(),
        "refresh_token": (dropbox_refresh_token or "").strip(),
        "app_key": (dropbox_app_key or "").strip(),
        "app_secret": (dropbox_app_secret or "").strip(),
        "_cached_access_token": None,
        "_cached_expires_at": None,
    }
    if auth["refresh_token"]:
        if not auth["app_key"] or not auth["app_secret"]:
            raise RuntimeError(
                "Dropbox refresh token requires both dropbox_app_key and dropbox_app_secret."
            )
        return auth
    if auth["access_token"]:
        return auth
    return None


def _skipped_upload_result(reason: str) -> dict[str, Any]:
    return {
        "ok": False,
        "status": "skipped",
        "error": reason,
    }


def _resolve_dropbox_access_token(dropbox_auth: dict[str, Any]) -> str:
    cached_token = dropbox_auth.get("_cached_access_token")
    cached_expires_at = dropbox_auth.get("_cached_expires_at")
    now_ts = _utc_now().timestamp()
    if cached_token and isinstance(cached_expires_at, (int, float)) and cached_expires_at > now_ts + 60:
        return str(cached_token)

    refresh_token = str(dropbox_auth.get("refresh_token") or "").strip()
    if refresh_token:
        payload = urlencode(
            {
                "grant_type": "refresh_token",
                "refresh_token": refresh_token,
                "client_id": str(dropbox_auth.get("app_key") or ""),
                "client_secret": str(dropbox_auth.get("app_secret") or ""),
            }
        ).encode("utf-8")
        req = request.Request(
            url="https://api.dropboxapi.com/oauth2/token",
            data=payload,
            method="POST",
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )
        with request.urlopen(req, timeout=20) as resp:
            raw = resp.read().decode("utf-8")
            token_payload = json.loads(raw) if raw else {}
            access_token = str(token_payload.get("access_token") or "").strip()
            expires_in = int(token_payload.get("expires_in") or 0)
            if not access_token:
                raise RuntimeError("Dropbox OAuth refresh succeeded but returned no access_token.")
            dropbox_auth["_cached_access_token"] = access_token
            dropbox_auth["_cached_expires_at"] = now_ts + max(0, expires_in)
            return access_token

    access_token = str(dropbox_auth.get("access_token") or "").strip()
    if not access_token:
        raise RuntimeError("Dropbox authentication is not configured.")
    return access_token


def _dropbox_error_details(raw_error: str) -> dict[str, Any]:
    text = (raw_error or "").strip()
    details: dict[str, Any] = {"raw": text}
    if not text:
        return details
    try:
        payload = json.loads(text)
    except Exception:
        if "required scope '" in text:
            scope = text.split("required scope '", 1)[1].split("'", 1)[0]
            details["reason"] = "missing_scope"
            details["friendly"] = f"Dropbox app is missing required scope {scope}. Enable it in the app console and mint a new token."
        return details

    summary = str(payload.get("error_summary") or "")
    error_obj = payload.get("error") or {}
    tag = error_obj.get(".tag") if isinstance(error_obj, dict) else None
    details.update({"payload": payload, "summary": summary, "tag": tag})

    if tag == "expired_access_token":
        details["reason"] = "expired_access_token"
        details["friendly"] = "Dropbox access token expired. Update the configured Dropbox token."
    elif tag == "invalid_access_token":
        details["reason"] = "invalid_access_token"
        details["friendly"] = "Dropbox access token is invalid. Update the configured Dropbox token."
    elif summary.startswith("path/not_found/"):
        details["reason"] = "path_not_found"
        details["friendly"] = "Dropbox file not found at the configured path."

    return details


def _github_push_file(
    local_path: str,
    repo_path: str,
    github_token: str,
    github_repo: str,
    github_branch: str = "main",
    commit_message: str | None = None,
) -> dict[str, Any]:
    """Push a single file to GitHub via the Contents API (no external deps)."""
    with open(local_path, "rb") as fh:
        content_b64 = base64.b64encode(fh.read()).decode("utf-8")

    msg = commit_message or f"Update session export: {Path(local_path).name}"
    api_url = f"https://api.github.com/repos/{github_repo}/contents/{repo_path}"
    headers = {
        "Authorization": f"Bearer {github_token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    # GET existing file to obtain SHA (needed for updates)
    sha: str | None = None
    try:
        get_req = request.Request(f"{api_url}?ref={github_branch}", headers=headers)
        with request.urlopen(get_req, timeout=15) as resp:
            sha = json.loads(resp.read().decode("utf-8")).get("sha")
    except urlerror.HTTPError as exc:
        if exc.code != 404:
            return {"ok": False, "path": repo_path, "error": exc.read().decode("utf-8"), "status": exc.code}

    payload: dict[str, Any] = {"message": msg, "content": content_b64, "branch": github_branch}
    if sha:
        payload["sha"] = sha

    put_req = request.Request(
        api_url,
        data=json.dumps(payload).encode("utf-8"),
        method="PUT",
        headers={**headers, "Content-Type": "application/json"},
    )
    try:
        with request.urlopen(put_req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return {"ok": True, "path": repo_path, "sha": data.get("content", {}).get("sha")}
    except urlerror.HTTPError as exc:
        return {"ok": False, "path": repo_path, "error": exc.read().decode("utf-8"), "status": exc.code}
    except Exception as exc:
        return {"ok": False, "path": repo_path, "error": str(exc)}


def _github_list_dir(
    github_token: str,
    github_repo: str,
    path: str,
    github_branch: str = "main",
) -> list[dict[str, Any]]:
    """List contents of a path in a GitHub repo via the Contents API. Returns [] on 404."""
    clean = path.strip("/")
    api_url = f"https://api.github.com/repos/{github_repo}/contents/{clean}"
    headers = {
        "Authorization": f"Bearer {github_token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    req = request.Request(f"{api_url}?ref={github_branch}", headers=headers)
    try:
        with request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urlerror.HTTPError as exc:
        if exc.code == 404:
            return []
        raise


def _github_download_url(url: str, github_token: str) -> bytes | None:
    """Download raw bytes from a GitHub download_url (handles private repos via auth)."""
    headers = {
        "Authorization": f"Bearer {github_token}",
        "Accept": "application/octet-stream",
    }
    req = request.Request(url, headers=headers)
    try:
        with request.urlopen(req, timeout=60) as resp:
            return resp.read()
    except urlerror.HTTPError as exc:
        if exc.code == 404:
            return None
        raise


def fetch_github_session_exports(
    github_token: str,
    github_repo: str,
    github_branch: str = "main",
    github_export_path: str = "session_exports",
    output_dir: str = "data/github_exports",
    limit: int = 3,
    expand_csv: bool = True,
) -> dict[str, Any]:
    """Download the latest N session export folders from the GitHub data repo.

    Files land in output_dir/{folder_timestamp}/ as .csv.gz and (if expand_csv)
    decompressed .csv so they can be read directly for analysis.
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    export_folder = github_export_path.strip("/")
    try:
        entries = _github_list_dir(github_token, github_repo, export_folder, github_branch)
    except Exception as exc:
        return {"ok": False, "error": str(exc), "fetched_count": 0, "folders": []}

    # Keep only directories, newest first (folder names are YYYY-MM-DD_HHMMSS)
    dirs = sorted(
        [e for e in entries if isinstance(e, dict) and e.get("type") == "dir"],
        key=lambda e: e.get("name", ""),
        reverse=True,
    )[:limit]

    if not dirs:
        return {
            "ok": True,
            "fetched_count": 0,
            "output_dir": output_dir,
            "folders": [],
            "note": f"No export folders found in {export_folder}",
        }

    fetched: list[dict[str, Any]] = []
    for folder_entry in dirs:
        folder_name = folder_entry["name"]
        folder_api_path = f"{export_folder}/{folder_name}"
        folder_out = output_path / folder_name
        folder_out.mkdir(parents=True, exist_ok=True)

        try:
            file_entries = _github_list_dir(
                github_token, github_repo, folder_api_path, github_branch
            )
        except Exception as exc:
            fetched.append({"folder": folder_name, "error": str(exc)})
            continue

        downloaded: list[str] = []
        errors: list[str] = []

        for f in file_entries:
            if not isinstance(f, dict) or f.get("type") != "file":
                continue
            fname = f["name"]
            download_url = f.get("download_url")
            if not download_url:
                continue

            local_file = folder_out / fname
            try:
                raw = _github_download_url(download_url, github_token)
                if raw is None:
                    errors.append(f"{fname}: not found")
                    continue
                local_file.write_bytes(raw)
                downloaded.append(fname)

                if expand_csv and fname.endswith(".csv.gz"):
                    csv_path = folder_out / fname[:-3]  # strip .gz
                    with gzip.open(str(local_file), "rb") as gz_in:
                        csv_path.write_bytes(gz_in.read())
                    downloaded.append(csv_path.name)
            except Exception as exc:
                errors.append(f"{fname}: {exc}")

        folder_info: dict[str, Any] = {
            "folder": folder_name,
            "local_dir": str(folder_out),
            "files": downloaded,
        }
        if errors:
            folder_info["errors"] = errors
        fetched.append(folder_info)

    logger.info(
        "Fetched %d GitHub session export folder(s) to %s",
        len(fetched),
        output_dir,
    )
    return {
        "ok": True,
        "fetched_count": len(fetched),
        "output_dir": output_dir,
        "folders": fetched,
    }


def _is_optional_latest_missing(item: dict[str, Any] | None, key: str) -> bool:
    if key not in _OPTIONAL_LATEST_KEYS:
        return False
    info = item or {}
    if info.get("ok"):
        return False
    details = info.get("error_details") or {}
    if details.get("reason") == "path_not_found":
        return True
    err_text = str(info.get("error") or "").lower()
    return "path/not_found" in err_text


def _db_iso(ts: datetime) -> str:
    """SQLite tables in this project store naive ISO timestamps."""
    return ts.replace(tzinfo=None).isoformat()


def _table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return {str(r[1]) for r in rows}


def _select_with_missing(conn: sqlite3.Connection, table: str, desired: list[str]) -> str:
    cols = _table_columns(conn, table)
    pieces: list[str] = []
    for col in desired:
        if col in cols:
            pieces.append(col)
        else:
            pieces.append(f"NULL AS {col}")
    return ", ".join(pieces)


def _aliased_select(table_alias: str, available_cols: set[str], mapping: list[tuple[str, str]]) -> str:
    pieces: list[str] = []
    for source, alias in mapping:
        if source in available_cols:
            pieces.append(f"{table_alias}.{source} AS {alias}")
        else:
            pieces.append(f"NULL AS {alias}")
    return ", ".join(pieces)


def _select_all(
    conn: sqlite3.Connection,
    query: str,
    params: tuple[Any, ...],
) -> tuple[list[str], list[tuple[Any, ...]]]:
    cur = conn.execute(query, params)
    columns = [d[0] for d in cur.description]
    rows = cur.fetchall()
    return columns, rows


def _latest_run_metadata(db_path: str) -> dict[str, Any] | None:
    conn = sqlite3.connect(db_path)
    try:
        cols = _table_columns(conn, "runs")
        if not cols:
            return None
        row = conn.execute(
            """SELECT run_id, started_at, app_version, strategy_version,
                      config_path, config_hash, dry_run, initial_mode,
                      default_order_size_usd, initial_paper_capital
               FROM runs ORDER BY started_at DESC LIMIT 1"""
        ).fetchone()
        if not row:
            return None
        return {
            "run_id": row[0],
            "started_at": row[1],
            "app_version": row[2],
            "strategy_version": row[3],
            "config_path": row[4],
            "config_hash": row[5],
            "dry_run": bool(row[6]),
            "mode": row[7],
            "order_size_usd": row[8],
            "paper_capital_total": row[9],
        }
    finally:
        conn.close()


def _sheet_from_rows(wb: Workbook, sheet_name: str, columns: list[str], rows: list[tuple[Any, ...]]) -> None:
    ws = wb.create_sheet(sheet_name)
    if not columns:
        ws.append(["no_data"])
        _style_header(ws, 1)
        _freeze(ws)
        _auto_width(ws)
        return
    ws.append(columns)
    _style_header(ws, len(columns))
    _freeze(ws)
    for row in rows:
        ws.append(list(row))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    _auto_width(ws)


def _session_export_folder_name(now_utc: datetime, label: str) -> str:
    return f"{now_utc.strftime('%Y-%m-%d_%H%M%S')}_{label}_session_export"


def _read_gzip_csv_rows(path: str) -> tuple[list[str], list[tuple[Any, ...]]]:
    with gzip.open(path, "rt", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        columns = next(reader, [])
        rows = [tuple(row) for row in reader]
    return columns, rows


def _session_column_meta(
    columns: list[str],
    compact_names: dict[str, str],
) -> list[tuple[str, str, str]]:
    reverse = {compact: verbose for verbose, compact in compact_names.items()}
    return [
        (str(column), reverse.get(str(column), str(column)), _verbose_header_title(reverse.get(str(column), str(column))))
        for column in columns
    ]


def _verbose_header_title(name: str) -> str:
    if name.endswith("_s"):
        return f"{_verbose_header_title(name[:-2])} (sec)"
    if name.endswith("_pct"):
        return f"{_verbose_header_title(name[:-4])} (%)"
    specials = {
        "id": "ID",
        "pnl": "P&L",
        "usd": "USD",
        "utc": "UTC",
        "json": "JSON",
        "btc": "BTC",
        "csv": "CSV",
        "gz": "GZ",
        "api": "API",
        "avg": "Avg",
        "max": "Max",
        "min": "Min",
    }
    words: list[str] = []
    for part in name.split("_"):
        lower = part.lower()
        words.append(specials.get(lower, part.capitalize()))
    title = " ".join(words)
    return (
        title.replace("Trade Id", "Trade ID")
        .replace("Decision Id", "Decision ID")
        .replace("Run Id", "Run ID")
        .replace("Window Id", "Window ID")
        .replace("App Version", "App Version")
        .replace("Strategy Type", "Strategy")
    )


def _coerce_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return None
    lowered = text.lower()
    if lowered in {"true", "false"}:
        return lowered == "true"
    try:
        if any(ch in text for ch in (".", "e", "E")):
            return float(text)
        return int(text)
    except ValueError:
        return text


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _session_lookup(columns: list[str], compact_names: dict[str, str]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for idx, (_, verbose, _) in enumerate(_session_column_meta(columns, compact_names)):
        raw = str(columns[idx])
        lookup[raw] = idx
        lookup[verbose] = idx
    return lookup


def _session_row_value(row: tuple[Any, ...], lookup: dict[str, int], *keys: str) -> Any:
    for key in keys:
        idx = lookup.get(key)
        if idx is None or idx >= len(row):
            continue
        value = row[idx]
        if value not in (None, ""):
            return value
    return None


def _session_cell_number_format(verbose_name: str) -> str | None:
    currency_names = {
        "paper_capital_total", "order_size_usd", "cost", "fee_total", "pnl",
        "entry_depth_side_usd", "opposite_depth_usd", "up_depth_usd", "down_depth_usd",
        "resolution_pnl",
    }
    price_names = {
        "entry_price", "entry_bid", "entry_ask", "target_price", "exit_price",
        "bid_at_exit", "ask_at_exit", "max_bid_seen", "min_bid_seen", "resolution_value",
    }
    share_names = {"shares", "shares_requested", "shares_filled"}
    if verbose_name in currency_names or verbose_name.endswith("_usd"):
        return '$#,##0.00;[Red]-$#,##0.00'
    if verbose_name in price_names:
        return '0.0000'
    if verbose_name in share_names:
        return '0.0000'
    if verbose_name.endswith("_s") or "score" in verbose_name or "ratio" in verbose_name or "velocity" in verbose_name:
        return '0.00'
    if verbose_name.endswith("_pct") or verbose_name.endswith("_delta"):
        return '0.00'
    return None


def _apply_session_value_style(cell: Any, verbose_name: str) -> None:
    value = cell.value
    if verbose_name == "pnl":
        numeric = _safe_float(value)
        if numeric is not None:
            if numeric > 0:
                cell.fill = _SESSION_POSITIVE_FILL
                cell.font = Font(color="0E7C45", bold=True)
            elif numeric < 0:
                cell.fill = _SESSION_NEGATIVE_FILL
                cell.font = Font(color="B42318", bold=True)
            else:
                cell.fill = _SESSION_NEUTRAL_FILL
    if verbose_name == "action":
        palette = {
            "TRADE": ("D9F7E8", "0E7C45"),
            "DRY_RUN_SIGNAL": ("E0F2FE", "1D4ED8"),
            "SUPPRESSED": ("FFF0D6", "B54708"),
        }
        colors = palette.get(str(value or "").upper())
        if colors:
            cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
            cell.font = Font(color=colors[1], bold=True)
    if verbose_name == "side":
        palette = {
            "UP": ("D9F7E8", "0E7C45"),
            "YES": ("D9F7E8", "0E7C45"),
            "DOWN": ("FDE2E2", "B42318"),
            "NO": ("FDE2E2", "B42318"),
        }
        colors = palette.get(str(value or "").upper())
        if colors:
            cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
            cell.font = Font(color=colors[1], bold=True)


def _build_session_data_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list[str],
    rows: list[tuple[Any, ...]],
    compact_names: dict[str, str],
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = "1D4ED8"
    if not columns:
        ws.append(["No session data"])
        _style_header(ws, 1)
        _freeze(ws)
        _auto_width(ws)
        return

    meta = _session_column_meta(columns, compact_names)
    ws.append([pretty for _, _, pretty in meta])
    for cell in ws[1]:
        cell.fill = _SESSION_ACCENT_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _SESSION_BORDER
    ws.row_dimensions[1].height = 28
    _freeze(ws)

    for row in rows:
        ws.append([_coerce_excel_value(value) for value in row])
        current_row = ws.max_row
        for col_idx, (_, verbose_name, _) in enumerate(meta, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = Alignment(vertical="top")
            cell.border = _SESSION_BORDER
            number_format = _session_cell_number_format(verbose_name)
            if number_format:
                cell.number_format = number_format
            _apply_session_value_style(cell, verbose_name)
        if current_row % 2 == 0:
            for col_idx in range(1, len(meta) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                if cell.fill.patternType is None:
                    cell.fill = _SESSION_NEUTRAL_FILL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(meta))}{max(ws.max_row, 1)}"
    ws.row_dimensions[2].height = 20
    _auto_width(ws, cap=30)


def _build_session_metrics(
    trade_columns: list[str],
    trade_rows: list[tuple[Any, ...]],
    signal_columns: list[str],
    signal_rows: list[tuple[Any, ...]],
) -> dict[str, Any]:
    trade_lookup = _session_lookup(trade_columns, _SESSION_TRADE_COMPACT_NAMES)
    signal_lookup = _session_lookup(signal_columns, _SESSION_SIGNAL_COMPACT_NAMES)

    total_pnl = 0.0
    wins = 0
    losses = 0
    open_trades = 0
    breakeven = 0
    best_trade: tuple[str, float] | None = None
    worst_trade: tuple[str, float] | None = None
    pnl_by_coin: defaultdict[str, float] = defaultdict(float)
    trade_counts_by_coin: Counter[str] = Counter()
    cumulative_pnl: list[tuple[int, float]] = []
    running_pnl = 0.0

    for index, row in enumerate(trade_rows, start=1):
        pnl_value = _safe_float(_session_row_value(row, trade_lookup, "pnl"))
        status_value = str(_session_row_value(row, trade_lookup, "status") or "").upper()
        exit_timestamp = _session_row_value(row, trade_lookup, "exit_timestamp")
        trade_id = str(_session_row_value(row, trade_lookup, "trade_id", "id") or index)
        coin = str(_session_row_value(row, trade_lookup, "coin", "c") or "N/A").upper()

        if pnl_value is not None:
            total_pnl += pnl_value
            running_pnl += pnl_value
        cumulative_pnl.append((index, running_pnl))
        trade_counts_by_coin[coin] += 1
        pnl_by_coin[coin] += pnl_value or 0.0

        if status_value in {"OPEN", "PENDING"} or (not exit_timestamp and pnl_value in (None, 0.0)):
            open_trades += 1
            continue
        if pnl_value is None or pnl_value == 0:
            breakeven += 1
            continue
        if pnl_value > 0:
            wins += 1
        else:
            losses += 1
        if best_trade is None or pnl_value > best_trade[1]:
            best_trade = (trade_id, pnl_value)
        if worst_trade is None or pnl_value < worst_trade[1]:
            worst_trade = (trade_id, pnl_value)

    action_counts: Counter[str] = Counter()
    for row in signal_rows:
        action = str(_session_row_value(row, signal_lookup, "action", "act") or "UNKNOWN").upper()
        action_counts[action] += 1

    closed_trades = wins + losses + breakeven
    win_rate = (wins / closed_trades * 100.0) if closed_trades else 0.0
    avg_pnl = (total_pnl / len(trade_rows)) if trade_rows else 0.0
    best_coin = max(pnl_by_coin.items(), key=lambda item: item[1], default=("N/A", 0.0))

    return {
        "total_pnl": total_pnl,
        "wins": wins,
        "losses": losses,
        "open_trades": open_trades,
        "breakeven": breakeven,
        "closed_trades": closed_trades,
        "win_rate": win_rate,
        "avg_pnl": avg_pnl,
        "best_trade": best_trade,
        "worst_trade": worst_trade,
        "best_coin": best_coin,
        "coin_count": len([coin for coin in trade_counts_by_coin if coin and coin != "N/A"]),
        "trade_count": len(trade_rows),
        "signal_count": len(signal_rows),
        "pnl_by_coin": sorted(pnl_by_coin.items(), key=lambda item: item[1], reverse=True),
        "trade_counts_by_coin": sorted(trade_counts_by_coin.items(), key=lambda item: item[1], reverse=True),
        "action_counts": sorted(action_counts.items(), key=lambda item: (-item[1], item[0])),
        "cumulative_pnl": cumulative_pnl or [(1, 0.0)],
    }


def _write_session_card(
    ws: Any,
    start_col: int,
    start_row: int,
    end_col: int,
    end_row: int,
    title: str,
    value: str,
    fill: PatternFill,
) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = f"{title}\n{value}"
    cell.fill = fill
    cell.font = Font(color="0F172A", bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _SESSION_BORDER
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for item in row:
            item.border = _SESSION_BORDER


def _write_session_table(
    ws: Any,
    *,
    title: str,
    start_row: int,
    start_col: int,
    items: list[tuple[str, Any]],
) -> None:
    end_col = start_col + 2
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    header = ws.cell(row=start_row, column=start_col)
    header.value = title
    header.fill = _SESSION_TITLE_FILL
    header.font = Font(color="FFFFFF", bold=True, size=12)
    header.alignment = Alignment(horizontal="left", vertical="center")
    header.border = _SESSION_BORDER
    for idx, (label, value) in enumerate(items, start=1):
        row = start_row + idx
        label_cell = ws.cell(row=row, column=start_col, value=label)
        label_cell.fill = _SESSION_SUBHEADER_FILL
        label_cell.font = Font(bold=True, color="0F172A")
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = _SESSION_BORDER
        value_cell = ws.cell(row=row, column=start_col + 1, value=value)
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell.border = _SESSION_BORDER
        value_cell.fill = _SESSION_PANEL_FILL
        value_cell.number_format = '$#,##0.00;[Red]-$#,##0.00' if isinstance(value, (int, float)) and "P&L" in label else "General"


def _build_session_chart_data_sheet(wb: Workbook, metrics: dict[str, Any]) -> Any:
    ws = wb.create_sheet("Chart Data")
    ws.sheet_state = "hidden"

    outcomes = [
        ("Wins", metrics["wins"]),
        ("Losses", metrics["losses"]),
        ("Open", metrics["open_trades"]),
        ("Breakeven", metrics["breakeven"]),
    ]
    ws["A1"] = "Trade Outcome"
    ws["B1"] = "Count"
    for row_idx, (label, value) in enumerate(outcomes, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)

    signal_actions = metrics["action_counts"] or [("NO_SIGNALS", 0)]
    ws["D1"] = "Signal Action"
    ws["E1"] = "Count"
    for row_idx, (label, value) in enumerate(signal_actions, start=2):
        ws.cell(row=row_idx, column=4, value=label)
        ws.cell(row=row_idx, column=5, value=value)

    coin_pnl = metrics["pnl_by_coin"][:8] or [("N/A", 0.0)]
    ws["G1"] = "Coin"
    ws["H1"] = "P&L"
    for row_idx, (label, value) in enumerate(coin_pnl, start=2):
        ws.cell(row=row_idx, column=7, value=label)
        ws.cell(row=row_idx, column=8, value=value)

    cumulative_pnl = metrics["cumulative_pnl"] or [(1, 0.0)]
    ws["J1"] = "Trade Number"
    ws["K1"] = "Cumulative P&L"
    for row_idx, (trade_no, value) in enumerate(cumulative_pnl, start=2):
        ws.cell(row=row_idx, column=10, value=trade_no)
        ws.cell(row=row_idx, column=11, value=value)

    return ws


def _add_session_charts(summary: Any, chart_data: Any, metrics: dict[str, Any]) -> None:
    outcome_chart = DoughnutChart()
    outcome_chart.title = "Trade Outcomes"
    outcome_chart.style = 10
    outcome_chart.holeSize = 55
    outcome_chart.height = 7
    outcome_chart.width = 8.5
    outcome_data = Reference(chart_data, min_col=2, min_row=1, max_row=5)
    outcome_labels = Reference(chart_data, min_col=1, min_row=2, max_row=5)
    outcome_chart.add_data(outcome_data, titles_from_data=True)
    outcome_chart.set_categories(outcome_labels)
    outcome_chart.varyColors = True
    outcome_chart.dataLabels = DataLabelList()
    outcome_chart.dataLabels.showVal = True
    outcome_chart.dataLabels.showPercent = True
    summary.add_chart(outcome_chart, "A28")

    signal_end_row = max(2, len(metrics["action_counts"]) + 1)
    signal_chart = BarChart()
    signal_chart.type = "bar"
    signal_chart.style = 11
    signal_chart.title = "Signal Actions"
    signal_chart.height = 7
    signal_chart.width = 8.5
    signal_chart.y_axis.title = "Action"
    signal_chart.x_axis.title = "Count"
    signal_data = Reference(chart_data, min_col=5, min_row=1, max_row=signal_end_row)
    signal_labels = Reference(chart_data, min_col=4, min_row=2, max_row=signal_end_row)
    signal_chart.add_data(signal_data, titles_from_data=True)
    signal_chart.set_categories(signal_labels)
    signal_chart.legend = None
    summary.add_chart(signal_chart, "F28")

    coin_end_row = max(2, min(9, len(metrics["pnl_by_coin"]) + 1))
    coin_chart = BarChart()
    coin_chart.style = 12
    coin_chart.title = "P&L by Coin"
    coin_chart.height = 7
    coin_chart.width = 8.5
    coin_chart.y_axis.title = "P&L"
    coin_data = Reference(chart_data, min_col=8, min_row=1, max_row=coin_end_row)
    coin_labels = Reference(chart_data, min_col=7, min_row=2, max_row=coin_end_row)
    coin_chart.add_data(coin_data, titles_from_data=True)
    coin_chart.set_categories(coin_labels)
    coin_chart.legend = None
    summary.add_chart(coin_chart, "A44")

    cumulative_end_row = max(2, len(metrics["cumulative_pnl"]) + 1)
    cumulative_chart = LineChart()
    cumulative_chart.style = 13
    cumulative_chart.title = "Cumulative Session P&L"
    cumulative_chart.height = 7
    cumulative_chart.width = 8.5
    cumulative_chart.y_axis.title = "P&L"
    cumulative_chart.x_axis.title = "Trade #"
    cumulative_data = Reference(chart_data, min_col=11, min_row=1, max_row=cumulative_end_row)
    cumulative_labels = Reference(chart_data, min_col=10, min_row=2, max_row=cumulative_end_row)
    cumulative_chart.add_data(cumulative_data, titles_from_data=True)
    cumulative_chart.set_categories(cumulative_labels)
    summary.add_chart(cumulative_chart, "F44")


def _build_session_summary_sheet(
    summary: Any,
    *,
    label: str,
    now_utc: datetime,
    reset_at: str,
    trades_path: str,
    signals_path: str,
    excel_filename: str,
    metrics: dict[str, Any],
) -> None:
    summary.sheet_view.zoomScale = 90
    summary.sheet_properties.tabColor = "8B5CF6"
    for col, width in {"A": 19, "B": 16, "C": 19, "D": 16, "E": 19, "F": 16, "G": 19, "H": 16}.items():
        summary.column_dimensions[col].width = width

    summary.merge_cells("A1:H2")
    title_cell = summary["A1"]
    title_cell.value = f"{label} Session Performance Report"
    title_cell.fill = _SESSION_TITLE_FILL
    title_cell.font = Font(color="FFFFFF", bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = _SESSION_BORDER

    summary.merge_cells("A3:H3")
    subtitle = summary["A3"]
    subtitle.value = (
        f"Readable session export generated {now_utc.strftime('%Y-%m-%d %H:%M:%S UTC')} "
        f"for the session beginning {reset_at}"
    )
    subtitle.fill = _SESSION_PANEL_FILL
    subtitle.font = Font(color="1E293B", italic=True)
    subtitle.alignment = Alignment(horizontal="center", vertical="center")
    subtitle.border = _SESSION_BORDER

    _write_session_card(summary, 1, 5, 2, 7, "Total P&L", f"${metrics['total_pnl']:,.2f}", _SESSION_POSITIVE_FILL if metrics["total_pnl"] >= 0 else _SESSION_NEGATIVE_FILL)
    _write_session_card(summary, 3, 5, 4, 7, "Win Rate", f"{metrics['win_rate']:.1f}%", _SESSION_SUBHEADER_FILL)
    _write_session_card(summary, 5, 5, 6, 7, "Trades", str(metrics["trade_count"]), _SESSION_PANEL_FILL)
    _write_session_card(summary, 7, 5, 8, 7, "Signals", str(metrics["signal_count"]), _SESSION_PANEL_FILL)
    _write_session_card(summary, 1, 9, 2, 11, "Wins", str(metrics["wins"]), _SESSION_POSITIVE_FILL)
    _write_session_card(summary, 3, 9, 4, 11, "Losses", str(metrics["losses"]), _SESSION_NEGATIVE_FILL)
    _write_session_card(summary, 5, 9, 6, 11, "Open", str(metrics["open_trades"]), _SESSION_WARNING_FILL)
    _write_session_card(summary, 7, 9, 8, 11, "Avg Trade P&L", f"${metrics['avg_pnl']:,.2f}", _SESSION_PANEL_FILL)

    details = [
        ("Session Start (UTC)", reset_at),
        ("Exported At (UTC)", now_utc.isoformat()),
        ("Coins Traded", metrics["coin_count"]),
        ("Best Coin", f"{metrics['best_coin'][0]} (${metrics['best_coin'][1]:,.2f})"),
        ("Best Trade", "N/A" if metrics["best_trade"] is None else f"Trade {metrics['best_trade'][0]} (${metrics['best_trade'][1]:,.2f})"),
        ("Worst Trade", "N/A" if metrics["worst_trade"] is None else f"Trade {metrics['worst_trade'][0]} (${metrics['worst_trade'][1]:,.2f})"),
    ]
    files = [
        ("Excel Workbook", excel_filename),
        ("Trades Archive", Path(trades_path).name),
        ("Signals Archive", Path(signals_path).name),
        ("Breakeven Trades", metrics["breakeven"]),
    ]
    signals = [(action.replace("_", " ").title(), count) for action, count in metrics["action_counts"][:6]]
    if not signals:
        signals = [("No Signals", 0)]

    _write_session_table(summary, title="Session Snapshot", start_row=14, start_col=1, items=details)
    _write_session_table(summary, title="Files Included", start_row=14, start_col=5, items=files)
    _write_session_table(summary, title="Signal Breakdown", start_row=22, start_col=5, items=signals)
    summary.freeze_panes = "A5"


def _build_session_excel_export(
    *,
    output_dir: str,
    label: str,
    now_utc: datetime,
    reset_at: str,
    trades_path: str,
    signals_path: str,
    trade_count: int,
    signal_count: int,
) -> str:
    date_stamp = now_utc.strftime("%Y-%m-%d")
    time_stamp = now_utc.strftime("%H%M%S")
    excel_filename = f"{date_stamp}_{time_stamp}_{label}_session_export.xlsx"
    excel_path = Path(output_dir) / excel_filename

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    trade_columns, trade_rows = _read_gzip_csv_rows(trades_path)
    signal_columns, signal_rows = _read_gzip_csv_rows(signals_path)
    metrics = _build_session_metrics(trade_columns, trade_rows, signal_columns, signal_rows)
    _build_session_summary_sheet(
        summary,
        label=label,
        now_utc=now_utc,
        reset_at=reset_at,
        trades_path=trades_path,
        signals_path=signals_path,
        excel_filename=excel_filename,
        metrics=metrics,
    )
    chart_data = _build_session_chart_data_sheet(wb, metrics)
    _add_session_charts(summary, chart_data, metrics)
    _build_session_data_sheet(wb, "Session Trades", trade_columns, trade_rows, _SESSION_TRADE_COMPACT_NAMES)
    _build_session_data_sheet(wb, "Session Signals", signal_columns, signal_rows, _SESSION_SIGNAL_COMPACT_NAMES)

    wb.save(excel_path)
    return str(excel_path)


def export_last_24h_excel(
    db_path: str,
    output_dir: str,
    label: str,
    now_utc: datetime | None = None,
) -> tuple[str, dict[str, int]]:
    now_utc = now_utc or _utc_now()
    since_utc = now_utc - timedelta(hours=24)
    since_iso = _db_iso(since_utc)

    conn = sqlite3.connect(db_path)
    try:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        pt_select = _select_with_missing(
            conn,
            "paper_trades",
            [
                "id",
                "timestamp",
                "run_id",
                "app_version",
                "strategy_version",
                "config_hash",
                "mode",
                "dry_run",
                "order_size_usd",
                "paper_capital_total",
                "market_slug",
                "window_id",
                "coin",
                "strategy_type",
                "signal_context",
                "signal_overlap_count",
                "side",
                "entry_price",
                "entry_bid",
                "entry_ask",
                "entry_spread",
                "target_price",
                "shares",
                "shares_requested",
                "shares_filled",
                "blocked_min_5_shares",
                "cost",
                "fee_total",
                "status",
                "exit_price",
                "pnl",
                "exit_reason",
                "exit_timestamp",
                "time_remaining_s",
                "bid_at_exit",
                "ask_at_exit",
                "exit_spread",
                "market_start_time",
                "market_end_time",
                "entry_depth_side_usd",
                "opposite_depth_usd",
                "depth_ratio",
                "max_bid_seen",
                "min_bid_seen",
                "time_to_max_bid_s",
                "time_to_min_bid_s",
                "first_profit_time_s",
                "scalp_hit",
                "high_confidence_hit",
                "signal_score",
                "score_velocity",
                "score_entry",
                "score_depth",
                "score_spread",
                "score_time",
                "score_balance",
                "target_delta",
                "hard_stop_delta",
                "mfe",
                "mae",
                "peak_net_pnl",
                "trough_net_pnl",
                "stall_exit_triggered",
                "hold_to_resolution",
                "loss_cut_threshold_pct",
                "loss_pct_at_exit",
                "favorable_excursion",
                "ever_profitable",
                "resolution_winner",
                "resolution_side_match",
                "resolution_value",
                "resolution_pnl",
                "would_have_won",
                "would_have_beaten_exit",
                "missed_upside_after_exit",
                "time_to_resolution_s",
                "recovered_after_exit",
            ],
        )
        pt_cols, pt_rows = _select_all(
            conn,
            f"""
            SELECT {pt_select}
            FROM paper_trades
            WHERE timestamp >= ?
            ORDER BY id DESC
            """,
            (since_iso,),
        )
        _sheet_from_rows(wb, "Paper Trades 24h", pt_cols, pt_rows)

        t_select = _select_with_missing(
            conn,
            "trades",
            [
                "id",
                "decision_id",
                "timestamp",
                "run_id",
                "app_version",
                "strategy_version",
                "config_hash",
                "mode",
                "dry_run",
                "order_size_usd",
                "paper_capital_total",
                "market_slug",
                "window_id",
                "coin",
                "strategy_type",
                "side",
                "up_price",
                "down_price",
                "entry_price",
                "target_price",
                "combined_cost",
                "fee_total",
                "shares",
                "shares_requested",
                "shares_filled",
                "blocked_min_5_shares",
                "status",
                "abort_cost",
                "error",
                "entry_order_submitted_at",
                "entry_filled_at",
                "entry_time_to_fill_s",
                "entry_limit_price",
                "entry_fill_price",
                "entry_slippage",
                "entry_fill_ratio",
                "exit_order_submitted_at",
                "exit_filled_at",
                "exit_limit_price",
                "exit_fill_price",
                "exit_slippage",
                "exit_reason",
                "exit_price",
                "pnl",
                "time_remaining_s",
                "bid_at_exit",
                "ask_at_exit",
                "exit_spread",
                "max_bid_seen",
                "min_bid_seen",
                "time_to_max_bid_s",
                "time_to_min_bid_s",
                "first_profit_time_s",
                "scalp_hit",
                "high_confidence_hit",
                "hold_to_resolution",
                "mfe",
                "mae",
                "peak_net_pnl",
                "trough_net_pnl",
                "stall_exit_triggered",
                "dynamic_loss_cut_pct",
                "loss_pct_at_exit",
                "favorable_excursion",
                "ever_profitable",
                "cancel_repost_count",
            ],
        )
        lt_cols, lt_rows = _select_all(
            conn,
            f"""
            SELECT {t_select}
            FROM trades
            WHERE timestamp >= ?
            ORDER BY id DESC
            """,
            (since_iso,),
        )
        _sheet_from_rows(wb, "Live Trades 24h", lt_cols, lt_rows)

        d_select = _select_with_missing(
            conn,
            "decisions",
            [
                "id",
                "timestamp",
                "run_id",
                "app_version",
                "strategy_version",
                "config_hash",
                "mode",
                "dry_run",
                "order_size_usd",
                "paper_capital_total",
                "market_slug",
                "window_id",
                "coin",
                "strategy_type",
                "signal_context",
                "signal_overlap_count",
                "suppressed_reason",
                "market_end_time",
                "market_start_time",
                "up_best_ask",
                "down_best_ask",
                "combined_cost",
                "btc_price",
                "coin_velocity_30s",
                "coin_velocity_60s",
                "up_depth_usd",
                "down_depth_usd",
                "time_remaining_s",
                "feed_count",
                "filter_passed",
                "filter_failed",
                "action",
                "reason",
                "source_prices_json",
                "source_ages_json",
                "source_dispersion_pct",
                "source_staleness_max_s",
                "source_staleness_avg_s",
            ],
        )
        d_cols, d_rows = _select_all(
            conn,
            f"""
            SELECT {d_select}
            FROM decisions
            WHERE timestamp >= ?
            ORDER BY id DESC
            """,
            (since_iso,),
        )
        _sheet_from_rows(wb, "Decisions 24h", d_cols, d_rows)

        sig_select = _select_with_missing(
            conn,
            "decisions",
            [
                "id",
                "timestamp",
                "run_id",
                "app_version",
                "strategy_version",
                "config_hash",
                "mode",
                "dry_run",
                "order_size_usd",
                "paper_capital_total",
                "market_slug",
                "window_id",
                "coin",
                "strategy_type",
                "action",
                "suppressed_reason",
                "reason",
                "signal_context",
                "signal_overlap_count",
                "entry_price",
                "target_price",
                "expected_profit_per_share",
                "signal_score",
                "score_velocity",
                "score_entry",
                "score_depth",
                "score_spread",
                "score_time",
                "score_balance",
                "time_remaining_s",
                "coin_velocity_30s",
                "coin_velocity_60s",
                "entry_bid",
                "entry_ask",
                "entry_spread",
                "entry_depth_side_usd",
                "opposite_depth_usd",
                "depth_ratio",
                "resignal_cooldown_s",
                "min_price_improvement",
                "last_signal_age_s",
                "filter_passed",
                "filter_failed",
                "source_prices_json",
                "source_ages_json",
                "source_dispersion_pct",
                "source_staleness_max_s",
                "source_staleness_avg_s",
            ],
        )
        sig_cols, sig_rows = _select_all(
            conn,
            f"""
            SELECT {sig_select}
            FROM decisions
            WHERE timestamp >= ?
              AND action IN ('DRY_RUN_SIGNAL', 'TRADE', 'SUPPRESSED')
            ORDER BY id DESC
            """,
            (since_iso,),
        )
        _sheet_from_rows(wb, "Signals 24h", sig_cols, sig_rows)

        summary = wb.create_sheet("Summary")
        summary_headers = ["Metric", "Value"]
        summary.append(summary_headers)
        _style_header(summary, len(summary_headers))
        summary_rows = [
            ("Label", label),
            ("Exported At (UTC)", now_utc.isoformat()),
            ("Window Start (UTC)", since_utc.isoformat()),
            ("Paper Trades (24h)", len(pt_rows)),
            ("Live Trades (24h)", len(lt_rows)),
            ("Decisions (24h)", len(d_rows)),
            ("Signals (24h)", len(sig_rows)),
        ]
        for r in summary_rows:
            summary.append(list(r))
        _freeze(summary)
        _auto_width(summary)

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        date_stamp = now_utc.strftime("%Y-%m-%d")
        path = str(Path(output_dir) / f"{date_stamp}_{label}_last24h.xlsx")
        wb.save(path)
        return path, {
            "paper_trades_24h": len(pt_rows),
            "live_trades_24h": len(lt_rows),
            "decisions_24h": len(d_rows),
            "signals_24h": len(sig_rows),
        }
    finally:
        conn.close()


def export_recent_trades_csv_gz(
    db_path: str,
    output_dir: str,
    label: str,
    limit: int,
    now_utc: datetime | None = None,
) -> tuple[str, int, int | None, int | None]:
    now_utc = now_utc or _utc_now()
    conn = sqlite3.connect(db_path)
    try:
        pt_cols = _table_columns(conn, "paper_trades")
        d_cols = _table_columns(conn, "decisions")

        pt_select = _aliased_select(
            "pt",
            pt_cols,
            [
                ("id", "trade_id"),
                ("timestamp", "timestamp"),
                ("run_id", "run_id"),
                ("app_version", "app_version"),
                ("strategy_version", "strategy_version"),
                ("config_hash", "config_hash"),
                ("mode", "mode"),
                ("dry_run", "dry_run"),
                ("order_size_usd", "order_size_usd"),
                ("paper_capital_total", "paper_capital_total"),
                ("market_slug", "market_slug"),
                ("window_id", "window_id"),
                ("coin", "coin"),
                ("strategy_type", "strategy_type"),
                ("signal_context", "signal_context"),
                ("signal_overlap_count", "signal_overlap_count"),
                ("side", "side"),
                ("entry_price", "entry_price"),
                ("entry_bid", "entry_bid"),
                ("entry_ask", "entry_ask"),
                ("entry_spread", "entry_spread"),
                ("target_price", "target_price"),
                ("shares", "shares"),
                ("shares_requested", "shares_requested"),
                ("shares_filled", "shares_filled"),
                ("blocked_min_5_shares", "blocked_min_5_shares"),
                ("cost", "cost"),
                ("fee_total", "fee_total"),
                ("status", "status"),
                ("exit_price", "exit_price"),
                ("pnl", "pnl"),
                ("exit_reason", "exit_reason"),
                ("exit_timestamp", "exit_timestamp"),
                ("time_remaining_s", "time_remaining_s"),
                ("bid_at_exit", "bid_at_exit"),
                ("ask_at_exit", "ask_at_exit"),
                ("exit_spread", "exit_spread"),
                ("market_start_time", "market_start_time"),
                ("market_end_time", "market_end_time"),
                ("entry_depth_side_usd", "entry_depth_side_usd"),
                ("opposite_depth_usd", "opposite_depth_usd"),
                ("depth_ratio", "depth_ratio"),
                ("signal_score", "signal_score"),
                ("score_velocity", "score_velocity"),
                ("score_entry", "score_entry"),
                ("score_depth", "score_depth"),
                ("score_spread", "score_spread"),
                ("score_time", "score_time"),
                ("score_balance", "score_balance"),
                ("target_delta", "target_delta"),
                ("hard_stop_delta", "hard_stop_delta"),
                ("max_bid_seen", "max_bid_seen"),
                ("min_bid_seen", "min_bid_seen"),
                ("time_to_max_bid_s", "time_to_max_bid_s"),
                ("time_to_min_bid_s", "time_to_min_bid_s"),
                ("first_profit_time_s", "first_profit_time_s"),
                ("scalp_hit", "scalp_hit"),
                ("high_confidence_hit", "high_confidence_hit"),
                ("mfe", "mfe"),
                ("mae", "mae"),
                ("peak_net_pnl", "peak_net_pnl"),
                ("trough_net_pnl", "trough_net_pnl"),
                ("stall_exit_triggered", "stall_exit_triggered"),
                ("hold_to_resolution", "hold_to_resolution"),
                ("loss_cut_threshold_pct", "loss_cut_threshold_pct"),
                ("loss_pct_at_exit", "loss_pct_at_exit"),
                ("favorable_excursion", "favorable_excursion"),
                ("ever_profitable", "ever_profitable"),
                ("resolution_winner", "resolution_winner"),
                ("resolution_side_match", "resolution_side_match"),
                ("resolution_value", "resolution_value"),
                ("resolution_pnl", "resolution_pnl"),
                ("would_have_won", "would_have_won"),
                ("would_have_beaten_exit", "would_have_beaten_exit"),
                ("missed_upside_after_exit", "missed_upside_after_exit"),
                ("time_to_resolution_s", "time_to_resolution_s"),
                ("recovered_after_exit", "recovered_after_exit"),
            ],
        )
        d_select = _aliased_select(
            "d",
            d_cols,
            [
                ("filter_passed", "filter_passed"),
                ("filter_failed", "filter_failed"),
                ("reason", "decision_reason"),
                ("research_cluster_id", "research_cluster_id"),
                ("research_cluster_n", "research_cluster_n"),
                ("research_cluster_win_pct", "research_cluster_win_pct"),
                ("research_cluster_avg_pnl", "research_cluster_avg_pnl"),
                ("research_policy_action", "research_policy_action"),
                ("coin_velocity_30s", "coin_velocity_30s"),
                ("coin_velocity_60s", "coin_velocity_60s"),
                ("up_depth_usd", "up_depth_usd"),
                ("down_depth_usd", "down_depth_usd"),
                ("time_remaining_s", "decision_time_remaining_s"),
                ("source_prices_json", "source_prices_json"),
                ("source_ages_json", "source_ages_json"),
                ("source_dispersion_pct", "source_dispersion_pct"),
                ("source_staleness_max_s", "source_staleness_max_s"),
                ("source_staleness_avg_s", "source_staleness_avg_s"),
            ],
        )
        columns, rows = _select_all(
            conn,
            f"""
            SELECT
                {pt_select},
                {d_select}
            FROM paper_trades pt
            LEFT JOIN decisions d ON d.id = pt.decision_id
            ORDER BY pt.id DESC
            LIMIT ?
            """,
            (limit,),
        )

        ids = [int(r[0]) for r in rows if r and r[0] is not None]
        newest = max(ids) if ids else None
        oldest = min(ids) if ids else None
        id_start = f"{oldest:06d}" if oldest is not None else "000000"
        id_end = f"{newest:06d}" if newest is not None else "000000"

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        date_stamp = now_utc.strftime("%Y-%m-%d")
        time_stamp = now_utc.strftime("%H%M%S")
        out_path = Path(output_dir) / f"{date_stamp}_{time_stamp}_{label}_trades_{id_start}-{id_end}.csv.gz"

        with gzip.open(out_path, "wt", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            compact_names = {
                "trade_id": "id", "timestamp": "ts", "run_id": "rid", "app_version": "av",
                "strategy_version": "sv", "config_hash": "ch", "mode": "md", "dry_run": "dr",
                "order_size_usd": "os", "paper_capital_total": "cap", "market_slug": "mkt",
                "window_id": "wid", "coin": "c", "strategy_type": "st", "signal_context": "ctx",
                "signal_overlap_count": "ov", "side": "sd", "entry_price": "ep", "entry_bid": "eb",
                "entry_ask": "ea", "entry_spread": "es", "target_price": "tp", "shares": "sh",
                "shares_requested": "srq", "shares_filled": "sfl", "blocked_min_5_shares": "b5",
                "cost": "cs", "fee_total": "fee", "status": "status", "exit_price": "xp",
                "pnl": "pnl", "exit_reason": "er", "exit_timestamp": "xt", "time_remaining_s": "tx",
                "bid_at_exit": "xb", "ask_at_exit": "xa", "exit_spread": "xs", "market_start_time": "ms",
                "market_end_time": "me", "entry_depth_side_usd": "eds", "opposite_depth_usd": "ods",
                "depth_ratio": "drt", "max_bid_seen": "maxb", "min_bid_seen": "minb",
                "time_to_max_bid_s": "ttmax", "time_to_min_bid_s": "ttmin", "first_profit_time_s": "tfp",
                "signal_score": "sg", "score_velocity": "sgv", "score_entry": "sge",
                "score_depth": "sgd", "score_spread": "sgs", "score_time": "sgt",
                "score_balance": "sgb", "target_delta": "td", "hard_stop_delta": "hsd",
                "scalp_hit": "sc", "high_confidence_hit": "hc", "filter_passed": "fp",
                "filter_failed": "ff", "decision_reason": "why", "coin_velocity_30s": "v30",
                "research_cluster_id": "rcid", "research_cluster_n": "rcn",
                "research_cluster_win_pct": "rcw", "research_cluster_avg_pnl": "rcp",
                "research_policy_action": "rpa",
                "coin_velocity_60s": "v60", "up_depth_usd": "du", "down_depth_usd": "dd",
                "decision_time_remaining_s": "te", "mfe": "mfe", "mae": "mae",
                "peak_net_pnl": "pnp", "trough_net_pnl": "tnp", "stall_exit_triggered": "sx",
                "hold_to_resolution": "hr", "loss_cut_threshold_pct": "lct",
                "loss_pct_at_exit": "lpx", "favorable_excursion": "fex",
                "ever_profitable": "evp", "resolution_winner": "rw",
                "resolution_side_match": "rsm", "resolution_value": "rv",
                "resolution_pnl": "rpn", "would_have_won": "whw",
                "would_have_beaten_exit": "wbe", "missed_upside_after_exit": "mux",
                "time_to_resolution_s": "ttr", "recovered_after_exit": "rae",
                "source_prices_json": "spj",
                "source_ages_json": "saj", "source_dispersion_pct": "sdp",
                "source_staleness_max_s": "ssx", "source_staleness_avg_s": "ssa",
            }
            writer.writerow([compact_names.get(col, col) for col in columns])
            writer.writerows(rows)

        return str(out_path), len(rows), oldest, newest
    finally:
        conn.close()


def export_recent_signals_csv_gz(
    db_path: str,
    output_dir: str,
    label: str,
    limit: int,
    now_utc: datetime | None = None,
) -> tuple[str, int, int | None, int | None]:
    now_utc = now_utc or _utc_now()
    conn = sqlite3.connect(db_path)
    try:
        d_cols = _table_columns(conn, "decisions")
        d_select = _aliased_select(
            "d",
            d_cols,
            [
                ("id", "decision_id"),
                ("timestamp", "timestamp"),
                ("run_id", "run_id"),
                ("app_version", "app_version"),
                ("strategy_version", "strategy_version"),
                ("config_hash", "config_hash"),
                ("mode", "mode"),
                ("dry_run", "dry_run"),
                ("order_size_usd", "order_size_usd"),
                ("paper_capital_total", "paper_capital_total"),
                ("market_slug", "market_slug"),
                ("window_id", "window_id"),
                ("coin", "coin"),
                ("strategy_type", "strategy_type"),
                ("action", "action"),
                ("suppressed_reason", "suppressed_reason"),
                ("reason", "reason"),
                ("signal_context", "signal_context"),
                ("signal_overlap_count", "signal_overlap_count"),
                ("entry_price", "entry_price"),
                ("target_price", "target_price"),
                ("expected_profit_per_share", "expected_profit_per_share"),
                ("signal_score", "signal_score"),
                ("score_velocity", "score_velocity"),
                ("score_entry", "score_entry"),
                ("score_depth", "score_depth"),
                ("score_spread", "score_spread"),
                ("score_time", "score_time"),
                ("score_balance", "score_balance"),
                ("time_remaining_s", "time_remaining_s"),
                ("coin_velocity_30s", "coin_velocity_30s"),
                ("coin_velocity_60s", "coin_velocity_60s"),
                ("entry_bid", "entry_bid"),
                ("entry_ask", "entry_ask"),
                ("entry_spread", "entry_spread"),
                ("entry_depth_side_usd", "entry_depth_side_usd"),
                ("opposite_depth_usd", "opposite_depth_usd"),
                ("depth_ratio", "depth_ratio"),
                ("resignal_cooldown_s", "resignal_cooldown_s"),
                ("min_price_improvement", "min_price_improvement"),
                ("last_signal_age_s", "last_signal_age_s"),
                ("research_cluster_id", "research_cluster_id"),
                ("research_cluster_n", "research_cluster_n"),
                ("research_cluster_win_pct", "research_cluster_win_pct"),
                ("research_cluster_avg_pnl", "research_cluster_avg_pnl"),
                ("research_policy_action", "research_policy_action"),
                ("filter_passed", "filter_passed"),
                ("filter_failed", "filter_failed"),
                ("source_prices_json", "source_prices_json"),
                ("source_ages_json", "source_ages_json"),
                ("source_dispersion_pct", "source_dispersion_pct"),
                ("source_staleness_max_s", "source_staleness_max_s"),
                ("source_staleness_avg_s", "source_staleness_avg_s"),
            ],
        )
        columns, rows = _select_all(
            conn,
            f"""
            SELECT {d_select}
            FROM decisions d
            WHERE d.action IN ('DRY_RUN_SIGNAL', 'TRADE', 'SUPPRESSED')
            ORDER BY d.id DESC
            LIMIT ?
            """,
            (limit,),
        )

        ids = [int(r[0]) for r in rows if r and r[0] is not None]
        newest = max(ids) if ids else None
        oldest = min(ids) if ids else None
        id_start = f"{oldest:06d}" if oldest is not None else "000000"
        id_end = f"{newest:06d}" if newest is not None else "000000"

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        date_stamp = now_utc.strftime("%Y-%m-%d")
        time_stamp = now_utc.strftime("%H%M%S")
        out_path = Path(output_dir) / f"{date_stamp}_{time_stamp}_{label}_signals_{id_start}-{id_end}.csv.gz"

        with gzip.open(out_path, "wt", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            compact_names = {
                "decision_id": "id", "timestamp": "ts", "run_id": "rid", "app_version": "av",
                "strategy_version": "sv", "config_hash": "ch", "mode": "md", "dry_run": "dr",
                "order_size_usd": "os", "paper_capital_total": "cap", "market_slug": "mkt",
                "window_id": "wid", "coin": "c", "strategy_type": "st", "action": "act",
                "suppressed_reason": "sup", "reason": "why", "signal_context": "ctx",
                "signal_overlap_count": "ov", "entry_price": "ep", "target_price": "tp",
                "expected_profit_per_share": "eps", "signal_score": "sg", "score_velocity": "sgv",
                "score_entry": "sge", "score_depth": "sgd", "score_spread": "sgs",
                "score_time": "sgt", "score_balance": "sgb", "time_remaining_s": "te",
                "coin_velocity_30s": "v30", "coin_velocity_60s": "v60", "entry_bid": "eb",
                "entry_ask": "ea", "entry_spread": "es", "entry_depth_side_usd": "eds",
                "opposite_depth_usd": "ods", "depth_ratio": "drt", "resignal_cooldown_s": "rcd",
                "min_price_improvement": "mpi", "last_signal_age_s": "lsa",
                "research_cluster_id": "rcid", "research_cluster_n": "rcn",
                "research_cluster_win_pct": "rcw", "research_cluster_avg_pnl": "rcp",
                "research_policy_action": "rpa",
                "filter_passed": "fp", "filter_failed": "ff", "source_prices_json": "spj",
                "source_ages_json": "saj", "source_dispersion_pct": "sdp",
                "source_staleness_max_s": "ssx", "source_staleness_avg_s": "ssa",
            }
            writer.writerow([compact_names.get(col, col) for col in columns])
            writer.writerows(rows)

        return str(out_path), len(rows), oldest, newest
    finally:
        conn.close()


def export_session_trades_csv_gz(
    db_path: str,
    output_dir: str,
    label: str,
    since_utc: str,
    now_utc: datetime | None = None,
) -> tuple[str, int, int | None, int | None]:
    """Export all paper trades since since_utc (ISO timestamp) as a gzipped CSV."""
    now_utc = now_utc or _utc_now()
    conn = sqlite3.connect(db_path)
    try:
        pt_cols = _table_columns(conn, "paper_trades")
        d_cols = _table_columns(conn, "decisions")

        pt_select = _aliased_select(
            "pt",
            pt_cols,
            [
                ("id", "trade_id"), ("timestamp", "timestamp"), ("run_id", "run_id"),
                ("app_version", "app_version"), ("strategy_version", "strategy_version"),
                ("config_hash", "config_hash"), ("mode", "mode"), ("dry_run", "dry_run"),
                ("order_size_usd", "order_size_usd"), ("paper_capital_total", "paper_capital_total"),
                ("market_slug", "market_slug"), ("window_id", "window_id"), ("coin", "coin"),
                ("strategy_type", "strategy_type"), ("signal_context", "signal_context"),
                ("signal_overlap_count", "signal_overlap_count"), ("side", "side"),
                ("entry_price", "entry_price"), ("entry_bid", "entry_bid"), ("entry_ask", "entry_ask"),
                ("entry_spread", "entry_spread"), ("target_price", "target_price"),
                ("shares", "shares"), ("shares_requested", "shares_requested"),
                ("shares_filled", "shares_filled"), ("blocked_min_5_shares", "blocked_min_5_shares"),
                ("cost", "cost"), ("fee_total", "fee_total"), ("status", "status"),
                ("exit_price", "exit_price"), ("pnl", "pnl"), ("exit_reason", "exit_reason"),
                ("exit_timestamp", "exit_timestamp"), ("time_remaining_s", "time_remaining_s"),
                ("bid_at_exit", "bid_at_exit"), ("ask_at_exit", "ask_at_exit"),
                ("exit_spread", "exit_spread"), ("market_start_time", "market_start_time"),
                ("market_end_time", "market_end_time"), ("entry_depth_side_usd", "entry_depth_side_usd"),
                ("opposite_depth_usd", "opposite_depth_usd"), ("depth_ratio", "depth_ratio"),
                ("signal_score", "signal_score"), ("score_velocity", "score_velocity"),
                ("score_entry", "score_entry"), ("score_depth", "score_depth"),
                ("score_spread", "score_spread"), ("score_time", "score_time"),
                ("score_balance", "score_balance"), ("target_delta", "target_delta"),
                ("hard_stop_delta", "hard_stop_delta"), ("max_bid_seen", "max_bid_seen"),
                ("min_bid_seen", "min_bid_seen"), ("time_to_max_bid_s", "time_to_max_bid_s"),
                ("time_to_min_bid_s", "time_to_min_bid_s"), ("first_profit_time_s", "first_profit_time_s"),
                ("scalp_hit", "scalp_hit"), ("high_confidence_hit", "high_confidence_hit"),
                ("mfe", "mfe"), ("mae", "mae"), ("peak_net_pnl", "peak_net_pnl"),
                ("trough_net_pnl", "trough_net_pnl"), ("stall_exit_triggered", "stall_exit_triggered"),
                ("hold_to_resolution", "hold_to_resolution"),
                ("loss_cut_threshold_pct", "loss_cut_threshold_pct"),
                ("loss_pct_at_exit", "loss_pct_at_exit"),
                ("favorable_excursion", "favorable_excursion"),
                ("ever_profitable", "ever_profitable"),
                ("resolution_winner", "resolution_winner"),
                ("resolution_side_match", "resolution_side_match"),
                ("resolution_value", "resolution_value"),
                ("resolution_pnl", "resolution_pnl"),
                ("would_have_won", "would_have_won"),
                ("would_have_beaten_exit", "would_have_beaten_exit"),
                ("missed_upside_after_exit", "missed_upside_after_exit"),
                ("time_to_resolution_s", "time_to_resolution_s"),
                ("recovered_after_exit", "recovered_after_exit"),
            ],
        )
        d_select = _aliased_select(
            "d",
            d_cols,
            [
                ("filter_passed", "filter_passed"), ("filter_failed", "filter_failed"),
                ("reason", "decision_reason"), ("coin_velocity_30s", "coin_velocity_30s"),
                ("research_cluster_id", "research_cluster_id"),
                ("research_cluster_n", "research_cluster_n"),
                ("research_cluster_win_pct", "research_cluster_win_pct"),
                ("research_cluster_avg_pnl", "research_cluster_avg_pnl"),
                ("research_policy_action", "research_policy_action"),
                ("coin_velocity_60s", "coin_velocity_60s"), ("up_depth_usd", "up_depth_usd"),
                ("down_depth_usd", "down_depth_usd"), ("time_remaining_s", "decision_time_remaining_s"),
                ("source_prices_json", "source_prices_json"),
                ("source_ages_json", "source_ages_json"),
                ("source_dispersion_pct", "source_dispersion_pct"),
                ("source_staleness_max_s", "source_staleness_max_s"),
                ("source_staleness_avg_s", "source_staleness_avg_s"),
            ],
        )
        columns, rows = _select_all(
            conn,
            f"""
            SELECT {pt_select}, {d_select}
            FROM paper_trades pt
            LEFT JOIN decisions d ON d.id = pt.decision_id
            WHERE pt.timestamp >= ?
            ORDER BY pt.id ASC
            """,
            (since_utc,),
        )

        ids = [int(r[0]) for r in rows if r and r[0] is not None]
        newest = max(ids) if ids else None
        oldest = min(ids) if ids else None

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        date_stamp = now_utc.strftime("%Y-%m-%d")
        time_stamp = now_utc.strftime("%H%M%S")
        out_path = Path(output_dir) / f"{date_stamp}_{time_stamp}_{label}_session_trades.csv.gz"

        with gzip.open(out_path, "wt", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            compact_names = {
                "trade_id": "id", "timestamp": "ts", "run_id": "rid", "app_version": "av",
                "strategy_version": "sv", "config_hash": "ch", "mode": "md", "dry_run": "dr",
                "order_size_usd": "os", "paper_capital_total": "cap", "market_slug": "mkt",
                "window_id": "wid", "coin": "c", "strategy_type": "st", "signal_context": "ctx",
                "signal_overlap_count": "ov", "side": "sd", "entry_price": "ep", "entry_bid": "eb",
                "entry_ask": "ea", "entry_spread": "es", "target_price": "tp", "shares": "sh",
                "shares_requested": "srq", "shares_filled": "sfl", "blocked_min_5_shares": "b5",
                "cost": "cs", "fee_total": "fee", "status": "status", "exit_price": "xp",
                "pnl": "pnl", "exit_reason": "er", "exit_timestamp": "xt", "time_remaining_s": "tx",
                "bid_at_exit": "xb", "ask_at_exit": "xa", "exit_spread": "xs", "market_start_time": "ms",
                "market_end_time": "me", "entry_depth_side_usd": "eds", "opposite_depth_usd": "ods",
                "depth_ratio": "drt", "max_bid_seen": "maxb", "min_bid_seen": "minb",
                "time_to_max_bid_s": "ttmax", "time_to_min_bid_s": "ttmin", "first_profit_time_s": "tfp",
                "signal_score": "sg", "score_velocity": "sgv", "score_entry": "sge",
                "score_depth": "sgd", "score_spread": "sgs", "score_time": "sgt",
                "score_balance": "sgb", "target_delta": "td", "hard_stop_delta": "hsd",
                "scalp_hit": "sc", "high_confidence_hit": "hc", "filter_passed": "fp",
                "filter_failed": "ff", "decision_reason": "why", "coin_velocity_30s": "v30",
                "research_cluster_id": "rcid", "research_cluster_n": "rcn",
                "research_cluster_win_pct": "rcw", "research_cluster_avg_pnl": "rcp",
                "research_policy_action": "rpa",
                "coin_velocity_60s": "v60", "up_depth_usd": "du", "down_depth_usd": "dd",
                "decision_time_remaining_s": "te", "mfe": "mfe", "mae": "mae",
                "peak_net_pnl": "pnp", "trough_net_pnl": "tnp", "stall_exit_triggered": "sx",
                "hold_to_resolution": "hr", "loss_cut_threshold_pct": "lct",
                "loss_pct_at_exit": "lpx", "favorable_excursion": "fex",
                "ever_profitable": "evp", "resolution_winner": "rw",
                "resolution_side_match": "rsm", "resolution_value": "rv",
                "resolution_pnl": "rpn", "would_have_won": "whw",
                "would_have_beaten_exit": "wbe", "missed_upside_after_exit": "mux",
                "time_to_resolution_s": "ttr", "recovered_after_exit": "rae",
                "source_prices_json": "spj",
                "source_ages_json": "saj", "source_dispersion_pct": "sdp",
                "source_staleness_max_s": "ssx", "source_staleness_avg_s": "ssa",
            }
            writer.writerow([compact_names.get(col, col) for col in columns])
            writer.writerows(rows)

        return str(out_path), len(rows), oldest, newest
    finally:
        conn.close()


def export_session_signals_csv_gz(
    db_path: str,
    output_dir: str,
    label: str,
    since_utc: str,
    now_utc: datetime | None = None,
) -> tuple[str, int, int | None, int | None]:
    """Export all signals since since_utc (ISO timestamp) as a gzipped CSV."""
    now_utc = now_utc or _utc_now()
    conn = sqlite3.connect(db_path)
    try:
        d_cols = _table_columns(conn, "decisions")
        d_select = _aliased_select(
            "d",
            d_cols,
            [
                ("id", "decision_id"), ("timestamp", "timestamp"), ("run_id", "run_id"),
                ("app_version", "app_version"), ("strategy_version", "strategy_version"),
                ("config_hash", "config_hash"), ("mode", "mode"), ("dry_run", "dry_run"),
                ("order_size_usd", "order_size_usd"), ("paper_capital_total", "paper_capital_total"),
                ("market_slug", "market_slug"), ("window_id", "window_id"), ("coin", "coin"),
                ("strategy_type", "strategy_type"), ("action", "action"),
                ("suppressed_reason", "suppressed_reason"), ("reason", "reason"),
                ("signal_context", "signal_context"), ("signal_overlap_count", "signal_overlap_count"),
                ("entry_price", "entry_price"), ("target_price", "target_price"),
                ("expected_profit_per_share", "expected_profit_per_share"),
                ("signal_score", "signal_score"), ("score_velocity", "score_velocity"),
                ("score_entry", "score_entry"), ("score_depth", "score_depth"),
                ("score_spread", "score_spread"), ("score_time", "score_time"),
                ("score_balance", "score_balance"), ("time_remaining_s", "time_remaining_s"),
                ("coin_velocity_30s", "coin_velocity_30s"), ("coin_velocity_60s", "coin_velocity_60s"),
                ("entry_bid", "entry_bid"), ("entry_ask", "entry_ask"), ("entry_spread", "entry_spread"),
                ("entry_depth_side_usd", "entry_depth_side_usd"),
                ("opposite_depth_usd", "opposite_depth_usd"), ("depth_ratio", "depth_ratio"),
                ("resignal_cooldown_s", "resignal_cooldown_s"),
                ("min_price_improvement", "min_price_improvement"),
                ("last_signal_age_s", "last_signal_age_s"),
                ("research_cluster_id", "research_cluster_id"),
                ("research_cluster_n", "research_cluster_n"),
                ("research_cluster_win_pct", "research_cluster_win_pct"),
                ("research_cluster_avg_pnl", "research_cluster_avg_pnl"),
                ("research_policy_action", "research_policy_action"),
                ("filter_passed", "filter_passed"), ("filter_failed", "filter_failed"),
                ("source_prices_json", "source_prices_json"),
                ("source_ages_json", "source_ages_json"),
                ("source_dispersion_pct", "source_dispersion_pct"),
                ("source_staleness_max_s", "source_staleness_max_s"),
                ("source_staleness_avg_s", "source_staleness_avg_s"),
            ],
        )

        columns, rows = _select_all(
            conn,
            f"""
            SELECT {d_select}
            FROM decisions d
            WHERE d.action IN ('DRY_RUN_SIGNAL', 'TRADE', 'SUPPRESSED')
              AND d.timestamp >= ?
            ORDER BY d.id ASC
            """,
            (since_utc,),
        )

        ids = [int(r[0]) for r in rows if r and r[0] is not None]
        newest = max(ids) if ids else None
        oldest = min(ids) if ids else None

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        date_stamp = now_utc.strftime("%Y-%m-%d")
        time_stamp = now_utc.strftime("%H%M%S")
        out_path = Path(output_dir) / f"{date_stamp}_{time_stamp}_{label}_session_signals.csv.gz"

        with gzip.open(out_path, "wt", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            compact_names = {
                "decision_id": "id", "timestamp": "ts", "run_id": "rid", "app_version": "av",
                "strategy_version": "sv", "config_hash": "ch", "mode": "md", "dry_run": "dr",
                "order_size_usd": "os", "paper_capital_total": "cap", "market_slug": "mkt",
                "window_id": "wid", "coin": "c", "strategy_type": "st", "action": "act",
                "suppressed_reason": "sup", "reason": "why", "signal_context": "ctx",
                "signal_overlap_count": "ov", "entry_price": "ep", "target_price": "tp",
                "expected_profit_per_share": "eps", "signal_score": "sg", "score_velocity": "sgv",
                "score_entry": "sge", "score_depth": "sgd", "score_spread": "sgs",
                "score_time": "sgt", "score_balance": "sgb", "time_remaining_s": "te",
                "coin_velocity_30s": "v30", "coin_velocity_60s": "v60", "entry_bid": "eb",
                "entry_ask": "ea", "entry_spread": "es", "entry_depth_side_usd": "eds",
                "opposite_depth_usd": "ods", "depth_ratio": "drt", "resignal_cooldown_s": "rcd",
                "min_price_improvement": "mpi", "last_signal_age_s": "lsa",
                "research_cluster_id": "rcid", "research_cluster_n": "rcn",
                "research_cluster_win_pct": "rcw", "research_cluster_avg_pnl": "rcp",
                "research_policy_action": "rpa",
                "filter_passed": "fp", "filter_failed": "ff", "source_prices_json": "spj",
                "source_ages_json": "saj", "source_dispersion_pct": "sdp",
                "source_staleness_max_s": "ssx", "source_staleness_avg_s": "ssa",
            }
            writer.writerow([compact_names.get(col, col) for col in columns])
            writer.writerows(rows)

        return str(out_path), len(rows), oldest, newest
    finally:
        conn.close()


def _dropbox_upload_file(local_path: str, dropbox_path: str, dropbox_auth: dict[str, Any]) -> dict[str, Any]:
    with open(local_path, "rb") as fh:
        body = fh.read()

    try:
        token = _resolve_dropbox_access_token(dropbox_auth)
        req = request.Request(
            url="https://content.dropboxapi.com/2/files/upload",
            data=body,
            method="POST",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/octet-stream",
                "Dropbox-API-Arg": json.dumps(
                    {"path": dropbox_path, "mode": "overwrite", "autorename": False, "mute": True}
                ),
            },
        )
        with request.urlopen(req, timeout=300) as resp:
            raw = resp.read().decode("utf-8")
            payload = json.loads(raw) if raw else {}
            if resp.status < 200 or resp.status >= 300:
                return {
                    "ok": False,
                    "status": resp.status,
                    "error": f"Unexpected Dropbox upload status {resp.status}",
                    "path": dropbox_path,
                }
            return {
                "ok": True,
                "status": resp.status,
                "bytes": len(body),
                "path": dropbox_path,
                "payload": payload,
            }
    except urlerror.HTTPError as e:
        err_body = ""
        try:
            err_body = e.read().decode("utf-8")
        except Exception:
            err_body = str(e)
        return {
            "ok": False,
            "status": e.code,
            "error": err_body,
            "error_details": _dropbox_error_details(err_body),
            "path": dropbox_path,
        }
    except Exception as e:
        return {
            "ok": False,
            "status": None,
            "error": str(e),
            "error_details": _dropbox_error_details(str(e)),
            "path": dropbox_path,
        }


def _dropbox_get_metadata(dropbox_path: str, dropbox_auth: dict[str, Any]) -> dict[str, Any]:
    try:
        token = _resolve_dropbox_access_token(dropbox_auth)
        req = request.Request(
            url="https://api.dropboxapi.com/2/files/get_metadata",
            data=json.dumps({"path": dropbox_path}).encode("utf-8"),
            method="POST",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
        )
        with request.urlopen(req, timeout=20) as resp:
            raw = resp.read().decode("utf-8")
            payload = json.loads(raw) if raw else {}
            return {"exists": True, "status": resp.status, "payload": payload}
    except urlerror.HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8")
        except Exception:
            body = str(e)
        # Dropbox uses 409 for not_found
        if e.code == 409:
            return {
                "exists": False,
                "status": e.code,
                "error": body,
                "error_details": _dropbox_error_details(body),
            }
        return {
            "exists": False,
            "status": e.code,
            "error": body,
            "error_details": _dropbox_error_details(body),
        }
    except Exception as e:
        return {
            "exists": False,
            "status": None,
            "error": str(e),
            "error_details": _dropbox_error_details(str(e)),
        }


def _dropbox_download_file(dropbox_path: str, dropbox_auth: dict[str, Any], local_path: str) -> dict[str, Any]:
    try:
        token = _resolve_dropbox_access_token(dropbox_auth)
        req = request.Request(
            url="https://content.dropboxapi.com/2/files/download",
            data=b"",
            method="POST",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "text/plain; charset=utf-8",
                "Dropbox-API-Arg": json.dumps({"path": dropbox_path}),
            },
        )
        with request.urlopen(req, timeout=300) as resp:
            body = resp.read()
            p = Path(local_path)
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_bytes(body)
            return {"ok": True, "status": resp.status, "bytes": len(body), "path": str(p)}
    except urlerror.HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8")
        except Exception:
            body = str(e)
        return {
            "ok": False,
            "status": e.code,
            "error": body,
            "error_details": _dropbox_error_details(body),
            "path": str(local_path),
        }
    except Exception as e:
        return {
            "ok": False,
            "status": None,
            "error": str(e),
            "error_details": _dropbox_error_details(str(e)),
            "path": str(local_path),
        }


def _dropbox_create_or_get_shared_link(dropbox_path: str, dropbox_auth: dict[str, Any]) -> str | None:
    """Return a Dropbox shared link URL for dropbox_path, creating one if needed. Returns None on failure."""
    try:
        token = _resolve_dropbox_access_token(dropbox_auth)
    except Exception:
        return None

    create_req = request.Request(
        url="https://api.dropboxapi.com/2/sharing/create_shared_link_with_settings",
        data=json.dumps({"path": dropbox_path}).encode("utf-8"),
        method="POST",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
    )
    try:
        with request.urlopen(create_req, timeout=20) as resp:
            result = json.loads(resp.read().decode("utf-8"))
            return result.get("url")
    except urlerror.HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8")
        except Exception:
            pass
        if e.code != 409:
            return None
        # A link already exists — extract URL from the error metadata
        try:
            payload = json.loads(body)
            err_obj = payload.get("error") or {}
            err_meta = err_obj.get("metadata") if isinstance(err_obj, dict) else None
            if isinstance(err_meta, dict) and err_meta.get("url"):
                return err_meta["url"]
        except Exception:
            pass
        # Fall back to listing existing shared links
        list_req = request.Request(
            url="https://api.dropboxapi.com/2/sharing/list_shared_links",
            data=json.dumps({"path": dropbox_path, "direct_only": True}).encode("utf-8"),
            method="POST",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        )
        try:
            with request.urlopen(list_req, timeout=20) as resp:
                result = json.loads(resp.read().decode("utf-8"))
                links = result.get("links") or []
                return links[0].get("url") if links else None
        except Exception:
            return None
    except Exception:
        return None


def get_or_upload_excel_link(
    local_path: str,
    output_dir: str,
    label: str,
    dropbox_root: str = "/",
    dropbox_token: str | None = None,
    dropbox_refresh_token: str | None = None,
    dropbox_app_key: str | None = None,
    dropbox_app_secret: str | None = None,
) -> tuple[str | None, str | None]:
    """Upload local_path to Dropbox if needed, then return (url_or_path, error).

    Checks the archive index first — if the file was successfully uploaded
    previously, skips the re-upload and goes straight to the shared link.
    Returns (result, None) on success or (None, reason) on failure.
    """
    try:
        dropbox_auth = _build_dropbox_auth(
            dropbox_token=dropbox_token,
            dropbox_refresh_token=dropbox_refresh_token,
            dropbox_app_key=dropbox_app_key,
            dropbox_app_secret=dropbox_app_secret,
        )
        if not dropbox_auth:
            return None, "Dropbox not configured (no token or refresh-token set)"

        local_name = Path(local_path).name
        dbx_path: str | None = None

        # Check the archive index for a known Dropbox path for this file,
        # but only trust it if the corresponding upload actually succeeded.
        label_s = _safe_label(label)
        index_path = Path(output_dir) / f"{label_s}_latest_index.json"
        if index_path.exists():
            try:
                with open(index_path, "r", encoding="utf-8") as fh:
                    idx = json.load(fh)
                dbx_files = idx.get("dropbox_files") or {}
                dbx_uploads = idx.get("dropbox_uploads") or {}
                for key, path_val in dbx_files.items():
                    if path_val and Path(path_val).name == local_name:
                        if (dbx_uploads.get(key) or {}).get("ok"):
                            dbx_path = path_val
                        break
            except Exception:
                pass

        if not dbx_path:
            root = _normalize_dropbox_root(dropbox_root)
            dbx_path = f"{root}/latest/{local_name}"
            result = _dropbox_upload_file(local_path, dbx_path, dropbox_auth)
            if not result.get("ok"):
                err = str(result.get("error") or result.get("status") or "unknown error")
                details = result.get("error_details") or {}
                friendly = (details.get("friendly") or "").strip()
                reason = friendly or err
                logger.warning("Dropbox upload failed for %s: %s", local_path, err)
                return None, f"Dropbox upload failed: {reason}"

        url = _dropbox_create_or_get_shared_link(dbx_path, dropbox_auth)
        if url:
            return url, None
        # Shared link unavailable (scope may be missing) — return the raw Dropbox path
        logger.warning("Shared link unavailable for %s; returning raw Dropbox path", dbx_path)
        return dbx_path, None
    except Exception as exc:
        logger.warning("get_or_upload_excel_link failed for %s: %s", local_path, exc)
        return None, str(exc)


def sync_dropbox_latest_to_local(
    dropbox_token: str | None = None,
    dropbox_refresh_token: str | None = None,
    dropbox_app_key: str | None = None,
    dropbox_app_secret: str | None = None,
    dropbox_root: str = "/",
    output_dir: str = "data/dropbox_sync",
    label: str = "EDEC-BOT",
    expand_trades_csv: bool = True,
    expand_signals_csv: bool = True,
) -> dict[str, Any]:
    """Pull stable latest archive files from Dropbox into a local folder."""
    dropbox_auth = _build_dropbox_auth(
        dropbox_token=dropbox_token,
        dropbox_refresh_token=dropbox_refresh_token,
        dropbox_app_key=dropbox_app_key,
        dropbox_app_secret=dropbox_app_secret,
    )
    if not dropbox_auth:
        raise RuntimeError("Dropbox authentication is not configured.")

    label = _safe_label(label)
    root = _normalize_dropbox_root(dropbox_root)
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    latest_filenames = {
        "latest_last24h_xlsx": f"{label}_latest_last24h.xlsx",
        "latest_trades_csv_gz": f"{label}_latest_trades.csv.gz",
        "latest_signals_csv_gz": f"{label}_latest_signals.csv.gz",
        "latest_index_json": f"{label}_latest_index.json",
    }
    remote_candidates = _dropbox_latest_remote_candidates(root, latest_filenames, label=label)
    local = {
        "latest_last24h_xlsx": str(out / f"{label}_latest_last24h.xlsx"),
        "latest_trades_csv_gz": str(out / f"{label}_latest_trades.csv.gz"),
        "latest_signals_csv_gz": str(out / f"{label}_latest_signals.csv.gz"),
        "latest_index_json": str(out / f"{label}_latest_index.json"),
    }

    downloads: dict[str, Any] = {}
    for key, candidates in remote_candidates.items():
        attempts: list[dict[str, Any]] = []
        chosen: dict[str, Any] | None = None
        for remote_path in candidates:
            res = _dropbox_download_file(
                dropbox_path=remote_path,
                dropbox_auth=dropbox_auth,
                local_path=local[key],
            )
            attempts.append(
                {
                    "remote_path": remote_path,
                    "ok": bool(res.get("ok")),
                    "status": res.get("status"),
                    "error": res.get("error"),
                    "error_details": res.get("error_details"),
                }
            )
            if res.get("ok"):
                chosen = res
                chosen["remote_path"] = remote_path
                break
        if chosen is None:
            chosen = {
                "ok": False,
                "status": attempts[-1].get("status") if attempts else None,
                "error": attempts[-1].get("error") if attempts else "No Dropbox path candidates built",
                "error_details": attempts[-1].get("error_details") if attempts else None,
                "path": local[key],
                "remote_path": candidates[0] if candidates else None,
            }
        if _is_optional_latest_missing(chosen, key):
            chosen["optional_missing"] = True
            details = dict(chosen.get("error_details") or {})
            details.setdefault("friendly", "Optional latest signals file is not in Dropbox yet.")
            chosen["error_details"] = details
        chosen["attempts"] = attempts
        downloads[key] = chosen

    expanded_csv = None
    if expand_trades_csv and downloads["latest_trades_csv_gz"].get("ok"):
        gz_path = Path(local["latest_trades_csv_gz"])
        csv_path = gz_path.with_suffix("")  # .csv.gz -> .csv
        with gzip.open(gz_path, "rb") as f_in, open(csv_path, "wb") as f_out:
            shutil.copyfileobj(f_in, f_out)
        expanded_csv = str(csv_path)
    expanded_signals_csv = None
    if expand_signals_csv and downloads["latest_signals_csv_gz"].get("ok"):
        gz_path = Path(local["latest_signals_csv_gz"])
        csv_path = gz_path.with_suffix("")
        with gzip.open(gz_path, "rb") as f_in, open(csv_path, "wb") as f_out:
            shutil.copyfileobj(f_in, f_out)
        expanded_signals_csv = str(csv_path)

    required_keys = tuple(key for key in downloads.keys() if key not in _OPTIONAL_LATEST_KEYS)
    ok = all(bool(downloads[key].get("ok")) for key in required_keys)
    return {
        "ok": ok,
        "checked_at_utc": _utc_now().isoformat(),
        "output_dir": str(out),
        "downloads": downloads,
        "expanded_trades_csv": expanded_csv,
        "expanded_signals_csv": expanded_signals_csv,
    }


def _safe_label(label: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "-" for ch in label).strip("-_") or "EDEC-BOT"


def _normalize_dropbox_root(dropbox_root: str | None) -> str:
    """Normalize Dropbox root so generated file paths are always valid."""
    root = (dropbox_root or "/").strip()
    if len(root) >= 2 and root[0] == root[-1] and root[0] in ("'", '"'):
        root = root[1:-1].strip()
    root = root.replace("\\", "/")
    if not root:
        root = "/"
    if "/home/" in root and "dropbox.com" in root:
        root = root.split("/home/", 1)[1].split("?", 1)[0].split("#", 1)[0].strip("/")
        if root.startswith("Apps/"):
            parts = root.split("/", 2)
            root = "/" if len(parts) <= 2 else f"/{parts[2]}"
        else:
            root = f"/{root}" if root else "/"
    if root.startswith("https://") or root.startswith("http://"):
        root = "/"
    if not root.startswith("/"):
        root = f"/{root}"
    while "//" in root:
        root = root.replace("//", "/")
    normalized = root.rstrip("/")
    return normalized  # "" means Dropbox app root; callers format as f"{root}/subpath/..."


def _dropbox_latest_remote_candidates(
    root: str,
    latest_filenames: dict[str, str],
    label: str | None = None,
) -> dict[str, list[str]]:
    base = root.rstrip("/")
    candidate_dirs: list[str] = []
    if base.endswith("/latest"):
        candidate_dirs.extend([base, base[: -len("/latest")] or "/"])
    else:
        candidate_dirs.extend([f"{base}/latest", base])
    normalized_label = _safe_label(label or "")
    if normalized_label:
        label_base = f"{base}/{normalized_label}" if base else f"/{normalized_label}"
        candidate_dirs.extend([f"{label_base}/latest", label_base])
        if base != "/":
            candidate_dirs.extend([f"/{normalized_label}/latest", f"/{normalized_label}"])

    seen_dirs: set[str] = set()
    unique_dirs: list[str] = []
    for d in candidate_dirs:
        d = d.rstrip("/") or "/"
        if d not in seen_dirs:
            seen_dirs.add(d)
            unique_dirs.append(d)

    result: dict[str, list[str]] = {}
    for key, filename in latest_filenames.items():
        result[key] = [f"{d}/{filename}" if d != "/" else f"/{filename}" for d in unique_dirs]
    return result


def run_daily_archive(
    db_path: str = "data/decisions.db",
    output_dir: str = "data/exports",
    label: str = "EDEC-BOT",
    recent_limit: int = 100,
    dropbox_token: str | None = None,
    dropbox_refresh_token: str | None = None,
    dropbox_app_key: str | None = None,
    dropbox_app_secret: str | None = None,
    dropbox_root: str = "/",
) -> dict[str, Any]:
    now_utc = _utc_now()
    dropbox_auth = _build_dropbox_auth(
        dropbox_token=dropbox_token,
        dropbox_refresh_token=dropbox_refresh_token,
        dropbox_app_key=dropbox_app_key,
        dropbox_app_secret=dropbox_app_secret,
    )
    label = _safe_label(label)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    export_errors: list[str] = []
    counts: dict[str, int] = {
        "paper_trades_24h": 0,
        "live_trades_24h": 0,
        "decisions_24h": 0,
    }
    excel_path: str | None = None
    try:
        excel_path, counts = export_last_24h_excel(
            db_path=db_path,
            output_dir=output_dir,
            label=label,
            now_utc=now_utc,
        )
    except Exception as exc:
        logger.exception("24h Excel export failed; continuing with CSV exports")
        export_errors.append(f"excel_export_failed: {exc}")
    recent_path, recent_count, oldest_id, newest_id = export_recent_trades_csv_gz(
        db_path=db_path,
        output_dir=output_dir,
        label=label,
        limit=recent_limit,
        now_utc=now_utc,
    )
    recent_signals_path, recent_signals_count, oldest_signal_id, newest_signal_id = export_recent_signals_csv_gz(
        db_path=db_path,
        output_dir=output_dir,
        label=label,
        limit=recent_limit,
        now_utc=now_utc,
    )

    latest_excel = str(output_path / f"{label}_latest_last24h.xlsx")
    latest_trades = str(output_path / f"{label}_latest_trades.csv.gz")
    latest_signals = str(output_path / f"{label}_latest_signals.csv.gz")
    if excel_path and Path(excel_path).exists():
        shutil.copy2(excel_path, latest_excel)
    shutil.copy2(recent_path, latest_trades)
    shutil.copy2(recent_signals_path, latest_signals)
    run_meta = _latest_run_metadata(db_path)

    index_path = output_path / f"{label}_latest_index.json"
    index = {
        "label": label,
        "exported_at_utc": now_utc.isoformat(),
        "window_hours": 24,
        "recent_trades_limit": recent_limit,
        "recent_signals_limit": recent_limit,
        "row_counts": {
            **counts,
            "recent_trades_rows": recent_count,
            "recent_signals_rows": recent_signals_count,
        },
        "trade_id_range": {
            "oldest": oldest_id,
            "newest": newest_id,
        },
        "signal_id_range": {
            "oldest": oldest_signal_id,
            "newest": newest_signal_id,
        },
        "local_files": {
            "daily_last24h_xlsx": Path(excel_path).name if excel_path else None,
            "daily_recent_trades_csv_gz": Path(recent_path).name,
            "daily_recent_signals_csv_gz": Path(recent_signals_path).name,
            "latest_last24h_xlsx": Path(latest_excel).name if Path(latest_excel).exists() else None,
            "latest_trades_csv_gz": Path(latest_trades).name,
            "latest_signals_csv_gz": Path(latest_signals).name,
            "latest_index_json": index_path.name,
        },
        "latest_run": run_meta,
        "export_errors": export_errors,
        "dropbox_files": None,
        "dropbox_uploads": None,
    }
    index_path.write_text(json.dumps(index, indent=2), encoding="utf-8")

    if dropbox_auth:
        root = _normalize_dropbox_root(dropbox_root)
        has_excel_export = bool(excel_path and Path(excel_path).exists())
        dbx_paths = {
            "daily_recent_trades_csv_gz": f"{root}/daily-archives/{Path(recent_path).name}",
            "daily_recent_signals_csv_gz": f"{root}/daily-archives/{Path(recent_signals_path).name}",
            "latest_trades_csv_gz": f"{root}/latest/{Path(latest_trades).name}",
            "latest_signals_csv_gz": f"{root}/latest/{Path(latest_signals).name}",
            "latest_index_json": f"{root}/latest/{index_path.name}",
        }
        if has_excel_export:
            dbx_paths["daily_last24h_xlsx"] = f"{root}/daily-reports/{Path(excel_path).name}"
            dbx_paths["latest_last24h_xlsx"] = f"{root}/latest/{Path(latest_excel).name}"
        upload_results = {
            "daily_recent_trades_csv_gz": _dropbox_upload_file(
                recent_path, dbx_paths["daily_recent_trades_csv_gz"], dropbox_auth
            ),
            "daily_recent_signals_csv_gz": _dropbox_upload_file(
                recent_signals_path, dbx_paths["daily_recent_signals_csv_gz"], dropbox_auth
            ),
            "latest_trades_csv_gz": _dropbox_upload_file(
                latest_trades, dbx_paths["latest_trades_csv_gz"], dropbox_auth
            ),
            "latest_signals_csv_gz": _dropbox_upload_file(
                latest_signals, dbx_paths["latest_signals_csv_gz"], dropbox_auth
            ),
            "latest_index_json": _dropbox_upload_file(
                str(index_path), dbx_paths["latest_index_json"], dropbox_auth
            ),
        }
        if has_excel_export:
            upload_results["daily_last24h_xlsx"] = _dropbox_upload_file(
                excel_path, dbx_paths["daily_last24h_xlsx"], dropbox_auth
            )
            upload_results["latest_last24h_xlsx"] = _dropbox_upload_file(
                latest_excel, dbx_paths["latest_last24h_xlsx"], dropbox_auth
            )
        else:
            skip_reason = "Excel export unavailable in this run"
            upload_results["daily_last24h_xlsx"] = _skipped_upload_result(skip_reason)
            upload_results["latest_last24h_xlsx"] = _skipped_upload_result(skip_reason)
        index["dropbox_files"] = dbx_paths
        index["dropbox_uploads"] = upload_results
        index_path.write_text(json.dumps(index, indent=2), encoding="utf-8")

    logger.info("Daily archive export complete: %s", json.dumps(index["local_files"]))
    return {
        "excel_path": excel_path,
        "recent_path": recent_path,
        "recent_signals_path": recent_signals_path,
        "latest_excel": latest_excel,
        "latest_trades": latest_trades,
        "latest_signals": latest_signals,
        "index_path": str(index_path),
        "row_counts": index["row_counts"],
        "trade_id_range": index["trade_id_range"],
        "signal_id_range": index["signal_id_range"],
        "dropbox_files": index["dropbox_files"],
        "dropbox_uploads": index["dropbox_uploads"],
    }


def run_session_export(
    db_path: str = "data/decisions.db",
    output_dir: str = "data/exports",
    label: str = "EDEC-BOT",
    dropbox_token: str | None = None,
    dropbox_refresh_token: str | None = None,
    dropbox_app_key: str | None = None,
    dropbox_app_secret: str | None = None,
    dropbox_root: str = "/",
    github_token: str | None = None,
    github_repo: str | None = None,
    github_branch: str = "main",
    github_export_path: str = "session_exports",
) -> dict[str, Any]:
    """Export all trades + signals since last stats reset → Dropbox + GitHub."""
    now_utc = _utc_now()
    label = _safe_label(label)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    session_folder = _session_export_folder_name(now_utc, label)
    session_output_path = output_path / session_folder
    session_output_path.mkdir(parents=True, exist_ok=True)

    # Resolve reset_at from DB
    conn = sqlite3.connect(db_path)
    try:
        row = conn.execute("SELECT reset_at FROM paper_capital WHERE id = 1").fetchone()
        reset_at = row[0] if row and row[0] else "1970-01-01"
    finally:
        conn.close()

    trades_path, trade_count, oldest_id, newest_id = export_session_trades_csv_gz(
        db_path=db_path, output_dir=str(session_output_path), label=label, since_utc=reset_at, now_utc=now_utc,
    )
    signals_path, signal_count, oldest_sig, newest_sig = export_session_signals_csv_gz(
        db_path=db_path, output_dir=str(session_output_path), label=label, since_utc=reset_at, now_utc=now_utc,
    )
    excel_path = _build_session_excel_export(
        output_dir=str(session_output_path),
        label=label,
        now_utc=now_utc,
        reset_at=reset_at,
        trades_path=trades_path,
        signals_path=signals_path,
        trade_count=trade_count,
        signal_count=signal_count,
    )

    date_stamp = now_utc.strftime("%Y-%m-%d")
    time_stamp = now_utc.strftime("%H%M%S")
    index_filename = f"{date_stamp}_{time_stamp}_{label}_session_index.json"
    index_path = session_output_path / index_filename
    index: dict[str, Any] = {
        "label": label,
        "exported_at_utc": now_utc.isoformat(),
        "session_since_utc": reset_at,
        "session_folder": session_folder,
        "row_counts": {
            "session_trades_rows": trade_count,
            "session_signals_rows": signal_count,
        },
        "trade_id_range": {"oldest": oldest_id, "newest": newest_id},
        "signal_id_range": {"oldest": oldest_sig, "newest": newest_sig},
        "local_files": {
            "session_excel_xlsx": Path(excel_path).name,
            "session_trades_csv_gz": Path(trades_path).name,
            "session_signals_csv_gz": Path(signals_path).name,
            "session_index_json": index_filename,
        },
        "dropbox_uploads": None,
        "dropbox_error": None,
        "github_pushes": None,
    }
    index_path.write_text(json.dumps(index, indent=2), encoding="utf-8")

    result: dict[str, Any] = {
        "session_dir": str(session_output_path),
        "session_folder": session_folder,
        "excel_path": excel_path,
        "trades_path": trades_path,
        "signals_path": signals_path,
        "index_path": str(index_path),
        "trade_count": trade_count,
        "signal_count": signal_count,
        "session_since_utc": reset_at,
        "dropbox_uploads": None,
        "dropbox_error": None,
        "github_pushes": None,
    }

    dropbox_auth = None
    dropbox_error = None
    try:
        dropbox_auth = _build_dropbox_auth(
            dropbox_token=dropbox_token,
            dropbox_refresh_token=dropbox_refresh_token,
            dropbox_app_key=dropbox_app_key,
            dropbox_app_secret=dropbox_app_secret,
        )
    except Exception as exc:  # noqa: BLE001
        dropbox_error = str(exc)

    if dropbox_auth:
        try:
            root = _normalize_dropbox_root(dropbox_root)
            folder = f"{root}/session-exports/{session_folder}"
            uploads = {
                "session_excel_xlsx": _dropbox_upload_file(
                    excel_path, f"{folder}/{Path(excel_path).name}", dropbox_auth
                ),
                "session_trades_csv_gz": _dropbox_upload_file(
                    trades_path, f"{folder}/{Path(trades_path).name}", dropbox_auth
                ),
                "session_signals_csv_gz": _dropbox_upload_file(
                    signals_path, f"{folder}/{Path(signals_path).name}", dropbox_auth
                ),
                "session_index_json": _dropbox_upload_file(
                    str(index_path), f"{folder}/{index_filename}", dropbox_auth
                ),
            }
            index["dropbox_uploads"] = uploads
            index_path.write_text(json.dumps(index, indent=2), encoding="utf-8")
            result["dropbox_uploads"] = uploads
        except Exception as exc:  # noqa: BLE001
            dropbox_error = str(exc)
    elif any(
        str(value or "").strip()
        for value in (dropbox_token, dropbox_refresh_token, dropbox_app_key, dropbox_app_secret)
    ):
        dropbox_error = dropbox_error or "Dropbox is configured incompletely."

    if dropbox_error:
        result["dropbox_error"] = dropbox_error

    if github_token and github_repo:
        export_folder = github_export_path.strip("/")
        pushes: dict[str, Any] = {}
        for local_path, key in [
            (excel_path, "session_excel_xlsx"),
            (trades_path, "session_trades_csv_gz"),
            (signals_path, "session_signals_csv_gz"),
            (str(index_path), "session_index_json"),
        ]:
            repo_file_path = f"{export_folder}/{session_folder}/{Path(local_path).name}"
            pushes[key] = _github_push_file(
                local_path=local_path,
                repo_path=repo_file_path,
                github_token=github_token,
                github_repo=github_repo,
                github_branch=github_branch,
                commit_message=f"Session export {session_folder}: {Path(local_path).name}",
            )
        index["github_pushes"] = pushes
        index_path.write_text(json.dumps(index, indent=2), encoding="utf-8")
        result["github_pushes"] = pushes

    logger.info(
        "Session export complete: %d trades, %d signals since %s -> %s",
        trade_count, signal_count, reset_at, session_output_path,
    )
    return result


def latest_archive_paths(output_dir: str = "data/exports", label: str = "EDEC-BOT") -> dict[str, str]:
    label = _safe_label(label)
    base = Path(output_dir)
    return {
        "latest_excel": str(base / f"{label}_latest_last24h.xlsx"),
        "latest_trades": str(base / f"{label}_latest_trades.csv.gz"),
        "latest_signals": str(base / f"{label}_latest_signals.csv.gz"),
        "latest_index": str(base / f"{label}_latest_index.json"),
    }


def read_latest_index(output_dir: str = "data/exports", label: str = "EDEC-BOT") -> dict[str, Any] | None:
    paths = latest_archive_paths(output_dir=output_dir, label=label)
    p = Path(paths["latest_index"])
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def archive_health_snapshot(
    output_dir: str = "data/exports",
    label: str = "EDEC-BOT",
    dropbox_token: str | None = None,
    dropbox_refresh_token: str | None = None,
    dropbox_app_key: str | None = None,
    dropbox_app_secret: str | None = None,
    dropbox_root: str = "/",
) -> dict[str, Any]:
    label = _safe_label(label)
    dropbox_auth = _build_dropbox_auth(
        dropbox_token=dropbox_token,
        dropbox_refresh_token=dropbox_refresh_token,
        dropbox_app_key=dropbox_app_key,
        dropbox_app_secret=dropbox_app_secret,
    )
    local_paths = latest_archive_paths(output_dir=output_dir, label=label)
    index = read_latest_index(output_dir=output_dir, label=label)

    health: dict[str, Any] = {
        "label": label,
        "checked_at_utc": _utc_now().isoformat(),
        "index": index,
        "local": {
            "latest_excel_exists": Path(local_paths["latest_excel"]).exists(),
            "latest_trades_exists": Path(local_paths["latest_trades"]).exists(),
            "latest_signals_exists": Path(local_paths["latest_signals"]).exists(),
            "latest_index_exists": Path(local_paths["latest_index"]).exists(),
        },
        "dropbox_live": None,
    }

    if dropbox_auth:
        root = _normalize_dropbox_root(dropbox_root)
        latest_remote = {
            "latest_last24h_xlsx": f"{root}/latest/{label}_latest_last24h.xlsx",
            "latest_trades_csv_gz": f"{root}/latest/{label}_latest_trades.csv.gz",
            "latest_signals_csv_gz": f"{root}/latest/{label}_latest_signals.csv.gz",
            "latest_index_json": f"{root}/latest/{label}_latest_index.json",
        }
        files: dict[str, Any] = {}
        for key, p in latest_remote.items():
            files[key] = {"path": p, **_dropbox_get_metadata(p, dropbox_auth)}
            if _is_optional_latest_missing(files[key], key):
                files[key]["optional_missing"] = True
        required_keys = tuple(key for key in files.keys() if key not in _OPTIONAL_LATEST_KEYS)
        live_ok = all(bool(files[key].get("exists")) for key in required_keys)
        health["dropbox_live"] = {
            "enabled": True,
            "ok": live_ok,
            "files": files,
        }

    return health


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate 24h EDEC exports and optional Dropbox sync."
    )
    parser.add_argument("--db-path", default="data/decisions.db")
    parser.add_argument("--output-dir", default="data/exports")
    parser.add_argument("--label", default="EDEC-BOT")
    parser.add_argument("--recent-limit", type=int, default=100)
    parser.add_argument("--dropbox-token", default=os.getenv("EDEC_DROPBOX_TOKEN"))
    parser.add_argument("--dropbox-refresh-token", default=os.getenv("EDEC_DROPBOX_REFRESH_TOKEN"))
    parser.add_argument("--dropbox-app-key", default=os.getenv("EDEC_DROPBOX_APP_KEY"))
    parser.add_argument("--dropbox-app-secret", default=os.getenv("EDEC_DROPBOX_APP_SECRET"))
    parser.add_argument("--dropbox-root", default=os.getenv("EDEC_DROPBOX_ROOT", "/"))
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    result = run_daily_archive(
        db_path=args.db_path,
        output_dir=args.output_dir,
        label=args.label,
        recent_limit=args.recent_limit,
        dropbox_token=args.dropbox_token,
        dropbox_refresh_token=args.dropbox_refresh_token,
        dropbox_app_key=args.dropbox_app_key,
        dropbox_app_secret=args.dropbox_app_secret,
        dropbox_root=args.dropbox_root,
    )
    print(json.dumps(result, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
