"""Export SQLite data to color-coded Excel workbook with full analysis sheets."""

import csv
import logging
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------

# Row status fills
WIN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # soft green
LOSS_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # soft pink
OPEN_FILL  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # soft blue
ZEBRA_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # off-white

# Exit reason cell fills (applied to exit_reason column only)
REASON_FILLS = {
    "profit_target":   PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid"),  # medium green
    "high_confidence": PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid"),  # medium blue
    "loss_cut":        PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),  # orange
    "stall_exit":      PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),  # amber
    "near_close":      PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),  # yellow
    "dead_leg":        PatternFill(start_color="C9B1D9", end_color="C9B1D9", fill_type="solid"),  # lavender
    "resolution":      PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),  # gray
    "manual":          PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),  # light gray
}

ACTION_FILLS = {
    "TRADE": PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid"),
    "DRY_RUN_SIGNAL": PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid"),
    "SUPPRESSED": PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),
    "SKIP": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
}

# Header fills
HEADER_FILL  = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")  # navy
SUBHEAD_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # pale blue
SUMMARY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow

HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
SUBHEAD_FONT = Font(bold=True, size=10)

WIN_FONT  = Font(bold=True, color="375623")   # dark green text
LOSS_FONT = Font(bold=True, color="9C0006")   # dark red text

# Trade journal section-divider row
SECTION_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # dark navy
SECTION_FONT = Font(color="FFFFFF", bold=True, size=10)
LEARN_FILL   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # soft yellow


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def export_to_excel(db_path: str = "data/decisions.db",
                    output_dir: str = "data",
                    today_only: bool = False) -> str:
    """Generate a color-coded Excel workbook. Returns the file path."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    wb = Workbook()
    date_str = datetime.utcnow().strftime("%Y-%m-%d") if today_only else None

    _build_paper_trades_sheet(wb, conn, date_str)
    _build_run_metadata_sheet(wb, conn)
    _build_trade_journal_sheet(wb, conn, date_str, "closed_loss")
    _build_trade_journal_sheet(wb, conn, date_str, "closed_win")
    _build_exit_reason_sheet(wb, conn, date_str)
    _build_coin_performance_sheet(wb, conn, date_str)
    _build_strategy_breakdown_sheet(wb, conn, date_str)
    _build_time_remaining_sheet(wb, conn, date_str)
    _build_price_level_sheet(wb, conn, date_str)
    _build_daily_summary_sheet(wb, conn)
    _build_trades_sheet(wb, conn, date_str)
    _build_decisions_sheet(wb, conn, date_str)
    _build_signals_sheet(wb, conn, date_str)
    _build_filter_performance_sheet(wb, conn, date_str)
    _build_skipped_winners_sheet(wb, conn, date_str)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filepath = str(Path(output_dir) / f"edec_export_{ts}.xlsx")
    wb.save(filepath)
    conn.close()
    logger.info(f"Excel export saved: {filepath}")
    return filepath


def export_recent_to_excel(db_path: str = "data/decisions.db",
                           output_dir: str = "data",
                           limit: int = 50) -> str:
    """Generate a compact CSV export with the last N trades for AI analysis."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    wb = Workbook()
    _build_recent_trades_sheet(wb, conn, limit)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    ws = wb[f"Last {limit} Trades"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filepath = str(Path(output_dir) / f"edec_recent{limit}_{ts}.csv")
    latest = str(Path(output_dir) / f"edec_recent{limit}_latest.csv")

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    # Stable latest snapshot path for automation pipelines
    with open(latest, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    conn.close()
    logger.info(f"Recent {limit} trades CSV export saved: {filepath} (latest={latest})")
    return filepath


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _style_header(ws, num_cols: int, row: int = 1):
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18


def _auto_width(ws, cap: int = 42):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, cap)


def _pnl_cell(cell):
    """Color and bold a P&L cell based on sign."""
    try:
        val = float(cell.value)
        if val > 0:
            cell.fill = WIN_FILL
            cell.font = WIN_FONT
        elif val < 0:
            cell.fill = LOSS_FILL
            cell.font = LOSS_FONT
    except (TypeError, ValueError):
        pass


def _pnl_col(ws, col: int, start: int = 2):
    for row in ws.iter_rows(min_row=start, min_col=col, max_col=col):
        _pnl_cell(row[0])


def _reason_cell(cell):
    """Color an exit_reason cell using REASON_FILLS palette."""
    fill = REASON_FILLS.get(str(cell.value or "").lower())
    if fill:
        cell.fill = fill
        cell.font = Font(bold=True)


def _action_cell(cell):
    """Color an action cell for decision/signal sheets."""
    fill = ACTION_FILLS.get(str(cell.value or "").upper())
    if fill:
        cell.fill = fill
        cell.font = Font(bold=True)


def _zebra(ws, row: int, num_cols: int):
    if row % 2 == 0:
        for col in range(1, num_cols + 1):
            c = ws.cell(row=row, column=col)
            if not c.fill or c.fill.patternType is None:
                c.fill = ZEBRA_FILL


def _freeze(ws):
    ws.freeze_panes = "A2"


def _summary(ws, start_row: int, items: list):
    """Append a labeled summary block below data. items = [(label, value), ...]"""
    ws.cell(row=start_row, column=1, value="-- SUMMARY --").font = Font(bold=True, italic=True)
    for i, (label, value) in enumerate(items):
        r = start_row + 1 + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True)
        lc.fill = SUMMARY_FILL
        vc = ws.cell(row=r, column=2, value=value)
        vc.fill = SUMMARY_FILL
        _pnl_cell(vc)


def _build_run_metadata_sheet(wb, conn):
    ws = wb.create_sheet("Run Metadata")
    headers = [
        "Run ID", "Started At", "App Version", "Strategy Version",
        "Config Path", "Config Hash", "Dry Run", "Initial Mode",
        "Default Order Size $", "Initial Paper Capital $",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    rows = conn.execute("""
        SELECT run_id, started_at, app_version, strategy_version,
               config_path, config_hash, dry_run, initial_mode,
               default_order_size_usd, initial_paper_capital
        FROM runs
        ORDER BY started_at DESC
        LIMIT 50
    """).fetchall()
    for row in rows:
        ws.append(list(row))
    _auto_width(ws, cap=64)


def _win_rate(wins, losses):
    return round((wins or 0) / max((wins or 0) + (losses or 0), 1) * 100, 1)


# ---------------------------------------------------------------------------
# Trade journal computed columns
# ---------------------------------------------------------------------------

def _momentum_alignment(side: str, vel: float | None) -> str:
    """Did 30s velocity support the trade direction?"""
    if vel is None:
        return "No data"
    side = (side or "").lower()
    if side not in ("up", "down"):
        return "N/A (dual-leg)"
    vel_str = f"{vel:+.2f}%"
    supported = (side == "up" and vel > 0) or (side == "down" and vel < 0)
    return f"Supported ({vel_str})" if supported else f"Counter-trend ({vel_str})"


def _exit_assessment(reason: str, entry: float | None, bid: float | None,
                     exit_price: float | None, remaining: float | None) -> str:
    """Human-readable summary of how the trade exited."""
    r = (reason or "").lower()
    if r == "profit_target":
        return "Fee-adjusted profit exit"
    if r == "high_confidence":
        return "High bid — early profit captured"
    if r == "loss_cut":
        if entry and bid:
            pct = (bid - entry) / entry * 100
            return f"Stop loss @ {bid:.3f} (entry {entry:.3f}, {pct:+.1f}%)"
        return "Stop loss triggered"
    if r == "near_close":
        rem = f"{remaining:.0f}s" if remaining is not None else "≤30s"
        direction = "profit" if (exit_price or 0) > (entry or 0) else "loss"
        return f"Time-forced @ {rem} — {direction}"
    if r == "dead_leg":
        return "Dead leg: one side collapsed"
    if r == "resolution":
        return "Won at $1.00 payout" if (exit_price or 0) >= 0.99 else "Lost at $0.00 resolution"
    if r == "manual":
        return "Manually closed"
    return "—"


def _key_lesson(status: str, reason: str, side: str, vel_30s: float | None) -> str:
    """Key learning takeaway for the trade."""
    is_win = (status or "").lower() == "closed_win"
    r = (reason or "").lower()
    s = (side or "").lower()

    if is_win:
        if r == "resolution":
            return "Correct directional call — held to $1.00 payout. Entry analysis was accurate."
        if r == "profit_target":
            return "Good entry: bid rose enough for a net-positive exit before expiry."
        if r == "high_confidence":
            return "Bid surged early — profit locked in. Consider if holding to resolution would have yielded more."
        if r == "near_close":
            return "Late recovery near expiry. Position held correctly under time pressure."
        if r == "dead_leg":
            return "Arb completed — dead leg sold, surviving leg resolved in our favour."
        return "Profitable trade."

    # Losses — check momentum alignment
    aligned: bool | None = None
    if vel_30s is not None and s in ("up", "down"):
        aligned = (s == "up" and vel_30s > 0) or (s == "down" and vel_30s < 0)

    if r == "resolution":
        if aligned is False:
            return "Counter-trend entry — velocity warned against this direction. Review entry filters."
        return "Market resolved against position. No exit fired before expiry — consider tightening loss cut."
    if r == "loss_cut":
        if aligned is False:
            return "Entered against momentum. Loss cut protected capital — velocity was the warning sign."
        return "Momentum looked valid but reversed. Consider tracking mid-trade velocity changes."
    if r == "near_close":
        return "Bid never recovered. Consider lower entry price cap or stricter time-remaining filter."
    if r == "dead_leg":
        return "Second leg never dipped into range — one-sided exposure resolved against us."
    return "Position closed at a loss."


# ---------------------------------------------------------------------------
# Sheet: Recent N Trades (compact, AI-analysis-friendly)
# ---------------------------------------------------------------------------

def _build_recent_trades_sheet(wb, conn, limit: int):
    """Single compact sheet: last N trades with all key fields for AI analysis."""
    ws = wb.create_sheet(f"Last {limit} Trades")

    headers = [
        "id", "d", "t", "c", "st", "sd",
        "rid", "av", "sv", "ch", "md", "dr", "os", "cap", "wid", "ctx", "ov",
        "ep", "tp", "eb", "ea", "es", "sh", "srq", "sfl", "b5", "cs", "fee",
        "te", "ts", "v30", "v60", "du", "dd", "eds", "ods", "drt", "ma",
        "sg", "sgv", "sge", "sgd", "sgs", "sgt", "sgb", "td", "hsd",
        "fp", "ff", "why",
        "er", "xp", "xb", "xa", "xs", "tx", "hd",
        "pnl", "pp", "maxb", "minb", "ttmax", "ttmin", "tfp", "sc", "hc",
        "mfe", "mae", "pnp", "tnp", "sx",
        "status",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    rows = conn.execute("""
        SELECT
            pt.id,
            pt.timestamp,
            pt.coin,
            pt.strategy_type,
            pt.side,
            pt.run_id,
            pt.app_version,
            pt.strategy_version,
            pt.config_hash,
            pt.mode,
            pt.dry_run,
            pt.order_size_usd,
            pt.paper_capital_total,
            pt.window_id,
            pt.signal_context,
            pt.signal_overlap_count,
            pt.entry_price,
            pt.target_price,
            pt.entry_bid,
            pt.entry_ask,
            pt.entry_spread,
            pt.shares,
            pt.shares_requested,
            pt.shares_filled,
            pt.blocked_min_5_shares,
            pt.cost,
            pt.fee_total,
            pt.market_start_time,
            pt.exit_reason,
            pt.exit_price,
            pt.bid_at_exit,
            pt.ask_at_exit,
            pt.exit_spread,
            pt.pnl,
            pt.exit_timestamp,
            pt.time_remaining_s,
            pt.max_bid_seen,
            pt.min_bid_seen,
            pt.time_to_max_bid_s,
            pt.time_to_min_bid_s,
            pt.first_profit_time_s,
            pt.scalp_hit,
            pt.high_confidence_hit,
            pt.status,
            pt.entry_depth_side_usd,
            pt.opposite_depth_usd,
            pt.depth_ratio,
            pt.signal_score,
            pt.score_velocity,
            pt.score_entry,
            pt.score_depth,
            pt.score_spread,
            pt.score_time,
            pt.score_balance,
            pt.target_delta,
            pt.hard_stop_delta,
            d.filter_passed,
            d.filter_failed,
            d.coin_velocity_30s,
            d.coin_velocity_60s,
            d.up_depth_usd,
            d.down_depth_usd,
            d.time_remaining_s AS entry_remaining,
            d.reason           AS decision_reason,
            pt.mfe,
            pt.mae,
            pt.peak_net_pnl,
            pt.trough_net_pnl,
            pt.stall_exit_triggered
        FROM paper_trades pt
        LEFT JOIN (
            SELECT market_slug, strategy_type, MAX(id) AS best_id
            FROM decisions
            WHERE action != 'SKIP'
            GROUP BY market_slug, strategy_type
        ) top_d ON top_d.market_slug   = pt.market_slug
               AND top_d.strategy_type = pt.strategy_type
        LEFT JOIN decisions d ON d.id = top_d.best_id
        ORDER BY pt.id DESC
        LIMIT ?
    """, (limit,)).fetchall()

    num_cols = len(headers)
    for ri, r in enumerate(rows, start=2):
        ts_str = str(r[1] or "")
        date_part = ts_str[:10] if len(ts_str) >= 10 else ts_str
        time_part = ts_str[11:19] if len(ts_str) >= 19 else ""

        hold_s = None
        if r[43] and r[1]:
            try:
                t_in  = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
                t_out = datetime.fromisoformat(str(r[43]).replace("Z", "+00:00"))
                hold_s = round((t_out - t_in).total_seconds(), 1)
            except Exception:
                pass

        pnl_pct = None
        if r[25] and r[25] > 0 and r[42] is not None:
            pnl_pct = round(r[42] / r[25] * 100, 2)

        side = str(r[4] or "").lower()
        vel30 = r[55]
        if vel30 is None or side not in ("up", "down"):
            momentum_align = None
        else:
            aligned = (side == "up" and vel30 > 0) or (side == "down" and vel30 < 0)
            momentum_align = "aligned" if aligned else "counter"

        out = [
            r[0], date_part, time_part, r[2], r[3], r[4],
            r[5], r[6], r[7], r[8], r[9], int(bool(r[10])) if r[10] is not None else None,
            r[11], r[12], r[13], r[14], r[15],
            r[16], r[17], r[18], r[19], r[20], r[21], r[22], r[23], r[24], r[25], r[26],
            r[59], r[27],
            round(r[55], 4) if r[55] is not None else None,
            round(r[56], 4) if r[56] is not None else None,
            round(r[57], 2) if r[57] is not None else None,
            round(r[58], 2) if r[58] is not None else None,
            r[44], r[45], r[46], momentum_align,
            r[47], r[48], r[49], r[50], r[51], r[52], r[53], r[54],
            r[60], r[61], r[62],
            r[28], r[29], r[30], r[31], r[32], r[33], hold_s,
            r[42], pnl_pct, r[34], r[35], r[36], r[37], r[38],
            int(bool(r[39])) if r[39] is not None else None,
            int(bool(r[40])) if r[40] is not None else None,
            r[63], r[64], r[65], r[66], int(bool(r[67])) if r[67] is not None else None,
            r[41],
        ]
        ws.append(out)

        # Row color by status
        status_str = str(r[41] or "").lower()
        if status_str == "closed_win":
            row_fill = WIN_FILL
        elif status_str == "closed_loss":
            row_fill = LOSS_FILL
        elif status_str == "open":
            row_fill = OPEN_FILL
        elif ri % 2 == 0:
            row_fill = ZEBRA_FILL
        else:
            row_fill = None

        if row_fill:
            for col in range(1, num_cols + 1):
                ws.cell(row=ri, column=col).fill = row_fill

        _pnl_cell(ws.cell(row=ri, column=58))   # P&L $
        _pnl_cell(ws.cell(row=ri, column=59))   # P&L %
        _reason_cell(ws.cell(row=ri, column=54))  # Exit Reason

    # Summary block below data
    total_rows = len(rows)
    if total_rows:
        wins   = sum(1 for r in rows if str(r[41] or "").lower() == "closed_win")
        losses = sum(1 for r in rows if str(r[41] or "").lower() == "closed_loss")
        open_p = sum(1 for r in rows if str(r[41] or "").lower() == "open")
        total_pnl = sum(r[42] for r in rows if r[42] is not None)
        _summary(ws, total_rows + 3, [
            ("Trades shown", total_rows),
            ("Wins", wins),
            ("Losses", losses),
            ("Open", open_p),
            ("Win rate", f"{_win_rate(wins, losses):.1f}%"),
            ("Total P&L $", round(total_pnl, 4)),
        ])

    _auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"


# ---------------------------------------------------------------------------
# Sheets: Loss Journal / Win Journal
# ---------------------------------------------------------------------------

def _build_trade_journal_sheet(wb, conn, date_str: str | None, status: str):
    """Detailed per-trade journal with entry reasoning, market context, and lessons."""
    sheet_name = "Loss Journal" if status == "closed_loss" else "Win Journal"
    ws = wb.create_sheet(sheet_name)

    # Row 1 — section divider labels (merged cells)
    sections = [
        ("TRADE INFO",      1,  7),
        ("ENTRY DECISION",  8, 18),
        ("TRADE PROGRESS", 19, 22),
        ("RESULT",         23, 27),
        ("LEARNING",       28, 30),
        ("RUNTIME & PATH", 31, 43),
    ]
    for title, c_start, c_end in sections:
        ws.merge_cells(start_row=1, start_column=c_start,
                       end_row=1, end_column=c_end)
        cell = ws.cell(row=1, column=c_start, value=title)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 16

    # Row 2 — column headers
    headers = [
        # TRADE INFO (1-7)
        "#", "Date", "Entry Time", "Coin", "Strategy", "Side", "Market",
        # ENTRY DECISION (8-18)
        "Entry $", "Target $", "Shares", "Cost $",
        "Filters Passed", "Filters Failed",
        "Vel 30s %", "Vel 60s %", "Book Depth UP $", "Book Depth DOWN $",
        "Why Entered",
        # TRADE PROGRESS (19-22)
        "Time @ Entry (s)", "Time @ Exit (s)", "Hold Duration (s)", "Bid @ Exit $",
        # RESULT (23-27)
        "Exit Reason", "Exit Price $", "P&L $", "P&L %", "Fees $",
        # LEARNING (28-30)
        "Momentum Alignment", "Exit Assessment", "Key Lesson",
        # RUNTIME & PATH (31-43)
        "Run ID", "App Ver", "Strategy Ver", "Config Hash", "Mode", "Order Size $",
        "Signal Context", "Entry Spread $", "Exit Spread $", "Max Bid $", "Min Bid $",
        "First Profit Time (s)", "HC Hit",
    ]
    ws.append(headers)
    _style_header(ws, len(headers), row=2)
    ws.freeze_panes = "A3"

    extra = "AND pt.timestamp LIKE ?" if date_str else ""
    params = (status, f"{date_str}%") if date_str else (status,)

    rows = conn.execute(f"""
        SELECT
            pt.id,
            pt.timestamp,
            pt.coin,
            pt.strategy_type,
            pt.side,
            pt.market_slug,
            pt.entry_price,
            pt.target_price,
            pt.shares,
            pt.cost,
            pt.fee_total,
            pt.exit_reason,
            pt.exit_price,
            pt.pnl,
            pt.exit_timestamp,
            pt.time_remaining_s   AS exit_remaining,
            pt.bid_at_exit,
            pt.status,
            pt.run_id,
            pt.app_version,
            pt.strategy_version,
            pt.config_hash,
            pt.mode,
            pt.order_size_usd,
            pt.signal_context,
            pt.entry_spread,
            pt.exit_spread,
            pt.max_bid_seen,
            pt.min_bid_seen,
            pt.first_profit_time_s,
            pt.high_confidence_hit,
            d.filter_passed,
            d.filter_failed,
            d.coin_velocity_30s,
            d.coin_velocity_60s,
            d.up_depth_usd,
            d.down_depth_usd,
            d.time_remaining_s    AS entry_remaining,
            d.reason              AS decision_reason
        FROM paper_trades pt
        LEFT JOIN (
            SELECT market_slug, strategy_type, MAX(id) AS best_id
            FROM decisions
            WHERE action != 'SKIP'
            GROUP BY market_slug, strategy_type
        ) top_d ON top_d.market_slug   = pt.market_slug
               AND top_d.strategy_type = pt.strategy_type
        LEFT JOIN decisions d ON d.id = top_d.best_id
        WHERE pt.status = ? {extra}
        ORDER BY pt.timestamp DESC
    """, params).fetchall()

    row_fill = LOSS_FILL if status == "closed_loss" else WIN_FILL

    for ri, r in enumerate(rows, start=3):
        ts = str(r[1] or "")
        date_part = ts[:10] if len(ts) >= 10 else ts
        time_part = ts[11:19] if len(ts) >= 19 else ""

        # Hold duration
        hold_s = None
        if r[14] and r[1]:
            try:
                t_in  = datetime.fromisoformat(ts.replace("Z", "+00:00"))
                t_out = datetime.fromisoformat(str(r[14]).replace("Z", "+00:00"))
                hold_s = round((t_out - t_in).total_seconds(), 1)
            except Exception:
                pass

        # P&L %
        pnl_pct = None
        if r[9] and r[9] > 0 and r[13] is not None:
            pnl_pct = round(r[13] / r[9] * 100, 2)

        momentum = _momentum_alignment(str(r[4] or ""), r[33])
        exit_ass = _exit_assessment(str(r[11] or ""), r[6], r[16], r[12], r[15])
        lesson   = _key_lesson(str(r[17] or ""), str(r[11] or ""), str(r[4] or ""), r[33])

        out = [
            r[0],          # #
            date_part,     # Date
            time_part,     # Entry Time
            r[2],          # Coin
            r[3],          # Strategy
            r[4],          # Side
            r[5],          # Market
            r[6],          # Entry $
            r[7],          # Target $
            r[8],          # Shares
            r[9],          # Cost $
            r[31],         # Filters Passed
            r[32],         # Filters Failed
            round(r[33], 4) if r[33] is not None else None,  # Vel 30s %
            round(r[34], 4) if r[34] is not None else None,  # Vel 60s %
            round(r[35], 2) if r[35] is not None else None,  # Book Depth UP $
            round(r[36], 2) if r[36] is not None else None,  # Book Depth DOWN $
            r[38],         # Why Entered
            r[37],         # Time @ Entry (s)
            r[15],         # Time @ Exit (s)
            hold_s,        # Hold Duration (s)
            r[16],         # Bid @ Exit $
            r[11],         # Exit Reason
            r[12],         # Exit Price $
            r[13],         # P&L $
            pnl_pct,       # P&L %
            r[10],         # Fees $
            momentum,      # Momentum Alignment
            exit_ass,      # Exit Assessment
            lesson,        # Key Lesson
            r[18],         # Run ID
            r[19],         # App Ver
            r[20],         # Strategy Ver
            r[21],         # Config Hash
            r[22],         # Mode
            r[23],         # Order Size $
            r[24],         # Signal Context
            r[25],         # Entry Spread $
            r[26],         # Exit Spread $
            r[27],         # Max Bid $
            r[28],         # Min Bid $
            r[29],         # First Profit Time (s)
            "yes" if r[30] else "no",  # HC Hit
        ]
        ws.append(out)

        # Colour the whole row win/loss
        for col in range(1, len(headers) + 1):
            ws.cell(row=ri, column=col).fill = row_fill

        _pnl_cell(ws.cell(row=ri, column=25))   # P&L $
        _pnl_cell(ws.cell(row=ri, column=26))   # P&L %
        _reason_cell(ws.cell(row=ri, column=23))  # Exit Reason

        # Learning columns — yellow tint + wrap
        for col in (28, 29, 30):
            cell = ws.cell(row=ri, column=col)
            cell.fill = LEARN_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[ri].height = 50

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 1: Paper Trades
# ---------------------------------------------------------------------------

def _build_paper_trades_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Paper Trades")
    headers = [
        "ID", "Opened At", "Coin", "Market Slug", "Strategy", "Side",
        "Entry $", "Target $", "Shares", "Cost $", "Fees $",
        "Status", "Exit Reason", "Exit Price $", "Bid At Exit $",
        "P&L $", "P&L %", "Time Left At Exit (s)", "Hold Duration (s)",
        "Market End Time",
        "Run ID", "App Version", "Strategy Version", "Config Hash", "Mode", "Dry Run",
        "Order Size $", "Paper Capital $", "Window ID", "Signal Context", "Overlap Count",
        "Entry Bid $", "Entry Ask $", "Entry Spread $", "Exit Ask $", "Exit Spread $",
        "Entry Depth $", "Opposite Depth $", "Depth Ratio",
        "Max Bid Seen $", "Min Bid Seen $", "Time To Max Bid (s)", "Time To Min Bid (s)",
        "First Profit Time (s)", "Scalp Hit", "High-Confidence Hit",
        "Signal Score", "Score Velocity", "Score Entry", "Score Depth",
        "Score Spread", "Score Time", "Score Balance",
        "Target Delta", "Hard Stop Delta", "MFE", "MAE",
        "Peak Net P&L", "Trough Net P&L", "Stall Exit",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    where = "WHERE timestamp LIKE ?" if date_str else ""
    params = (f"{date_str}%",) if date_str else ()
    rows = conn.execute(f"""
        SELECT id, timestamp, coin, market_slug, strategy_type, side,
               entry_price, target_price, shares, cost, fee_total,
               status, exit_reason, exit_price, bid_at_exit, pnl,
               time_remaining_s, exit_timestamp, market_end_time,
               run_id, app_version, strategy_version, config_hash, mode, dry_run,
               order_size_usd, paper_capital_total, window_id, signal_context, signal_overlap_count,
               entry_bid, entry_ask, entry_spread, ask_at_exit, exit_spread,
               entry_depth_side_usd, opposite_depth_usd, depth_ratio,
               max_bid_seen, min_bid_seen, time_to_max_bid_s, time_to_min_bid_s,
               first_profit_time_s, scalp_hit, high_confidence_hit,
               signal_score, score_velocity, score_entry, score_depth,
               score_spread, score_time, score_balance,
               target_delta, hard_stop_delta, mfe, mae,
               peak_net_pnl, trough_net_pnl, stall_exit_triggered
        FROM paper_trades {where}
        ORDER BY id DESC LIMIT 10000
    """, params)

    for ri, row in enumerate(rows, start=2):
        r = list(row)
        cost, pnl = r[9] or 0, r[15]
        pnl_pct = round(pnl / cost * 100, 1) if (pnl is not None and cost > 0) else None
        # Hold duration
        hold_s = None
        if r[1] and r[17]:
            try:
                o = datetime.fromisoformat(str(r[1]).replace("Z", ""))
                e = datetime.fromisoformat(str(r[17]).replace("Z", ""))
                hold_s = round((e - o).total_seconds(), 1)
            except Exception:
                pass
        out = r[:16] + [pnl_pct, r[16], hold_s, r[18]] + r[19:]
        ws.append(out)

        status = str(r[11] or "").lower()
        row_fill = WIN_FILL if status == "closed_win" else (
                   LOSS_FILL if status == "closed_loss" else (
                   OPEN_FILL if status == "open" else None))
        if row_fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ri, column=col).fill = row_fill

        _reason_cell(ws.cell(row=ri, column=13))   # exit_reason col
        _pnl_cell(ws.cell(row=ri, column=16))       # P&L $
        _pnl_cell(ws.cell(row=ri, column=17))       # P&L %

    # Summary
    stats = conn.execute("""
        SELECT
            COUNT(*),
            SUM(CASE WHEN status='closed_win' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status='closed_loss' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status='open' THEN 1 ELSE 0 END),
            ROUND(SUM(COALESCE(pnl,0)), 4),
            ROUND(AVG(entry_price), 4),
            ROUND(AVG(CASE WHEN status IN ('closed_win','closed_loss') THEN exit_price END), 4),
            ROUND(AVG(CASE WHEN status='closed_win' AND cost>0 THEN pnl/cost*100 END), 2),
            ROUND(AVG(CASE WHEN status='closed_loss' AND cost>0 THEN pnl/cost*100 END), 2)
        FROM paper_trades
    """).fetchone()
    if stats:
        total, wins, losses, open_pos, total_pnl, avg_buy, avg_sell, avg_win_pct, avg_loss_pct = stats
        sr = ws.max_row + 2
        _summary(ws, sr, [
            ("Total Trades", total or 0),
            ("Wins", wins or 0),
            ("Losses", losses or 0),
            ("Open", open_pos or 0),
            ("Win Rate %", _win_rate(wins, losses)),
            ("Total P&L $", round(total_pnl or 0, 4)),
            ("Avg Entry Price", avg_buy or 0),
            ("Avg Exit Price (closed)", avg_sell or 0),
            ("Avg Win %", avg_win_pct or 0),
            ("Avg Loss %", avg_loss_pct or 0),
        ])

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 2: Exit Reason Analysis
# ---------------------------------------------------------------------------

def _build_exit_reason_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Exit Reason Analysis")
    headers = [
        "Exit Reason", "Trades", "Wins", "Losses", "Win Rate %",
        "Total P&L $", "Avg P&L $", "Avg P&L %",
        "Avg Time Remaining (s)", "Avg Entry $", "Avg Exit $",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    extra = "AND timestamp LIKE ?" if date_str else ""
    params = (f"{date_str}%",) if date_str else ()
    rows = conn.execute(f"""
        SELECT
            COALESCE(exit_reason, 'unknown') as reason,
            COUNT(*),
            SUM(CASE WHEN status='closed_win' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status='closed_loss' THEN 1 ELSE 0 END),
            ROUND(SUM(COALESCE(pnl,0)), 4) AS total_pnl,
            ROUND(AVG(COALESCE(pnl,0)), 4),
            ROUND(AVG(CASE WHEN cost>0 THEN pnl/cost*100 END), 2),
            ROUND(AVG(time_remaining_s), 1),
            ROUND(AVG(entry_price), 4),
            ROUND(AVG(exit_price), 4)
        FROM paper_trades
        WHERE status IN ('closed_win','closed_loss') {extra}
        GROUP BY reason ORDER BY 5 DESC
    """, params)

    for ri, row in enumerate(rows, start=2):
        r = list(row)
        wins, losses = r[2] or 0, r[3] or 0
        out = [r[0], r[1], wins, losses, _win_rate(wins, losses),
               r[4], r[5], r[6], r[7], r[8], r[9]]
        ws.append(out)
        _reason_cell(ws.cell(row=ri, column=1))
        _pnl_cell(ws.cell(row=ri, column=6))
        _pnl_cell(ws.cell(row=ri, column=7))
        _pnl_cell(ws.cell(row=ri, column=8))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 3: Coin Performance
# ---------------------------------------------------------------------------

def _build_coin_performance_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Coin Performance")
    headers = [
        "Coin", "Trades", "Wins", "Losses", "Open",
        "Win Rate %", "Total P&L $", "Avg P&L $", "Avg P&L %",
        "Avg Entry $", "Avg Exit $", "Best Trade $", "Worst Trade $",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    where = "WHERE timestamp LIKE ?" if date_str else ""
    params = (f"{date_str}%",) if date_str else ()
    rows = conn.execute(f"""
        SELECT
            coin,
            COUNT(*),
            SUM(CASE WHEN status='closed_win' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status='closed_loss' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status='open' THEN 1 ELSE 0 END),
            ROUND(SUM(COALESCE(pnl,0)), 4) AS total_pnl,
            ROUND(AVG(COALESCE(pnl,0)), 4),
            ROUND(AVG(CASE WHEN cost>0 THEN pnl/cost*100 END), 2),
            ROUND(AVG(entry_price), 4),
            ROUND(AVG(CASE WHEN status IN ('closed_win','closed_loss') THEN exit_price END), 4),
            ROUND(MAX(COALESCE(pnl,0)), 4),
            ROUND(MIN(COALESCE(pnl,0)), 4)
        FROM paper_trades {where}
        GROUP BY coin ORDER BY 6 DESC
    """, params)

    for ri, row in enumerate(rows, start=2):
        r = list(row)
        wins, losses = r[2] or 0, r[3] or 0
        out = [r[0], r[1], wins, losses, r[4], _win_rate(wins, losses),
               r[5], r[6], r[7], r[8], r[9], r[10], r[11]]
        ws.append(out)
        for col in (7, 8, 9, 12, 13):
            _pnl_cell(ws.cell(row=ri, column=col))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 4: Strategy Breakdown
# ---------------------------------------------------------------------------

def _build_strategy_breakdown_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Strategy Breakdown")
    headers = [
        "Strategy", "Exit Reason", "Trades", "Wins", "Losses",
        "Win Rate %", "Total P&L $", "Avg P&L $", "Avg P&L %",
        "Avg Entry $", "Avg Time Remaining (s)",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    extra = "AND timestamp LIKE ?" if date_str else ""
    params = (f"{date_str}%",) if date_str else ()
    rows = conn.execute(f"""
        SELECT
            strategy_type,
            COALESCE(exit_reason, 'unknown'),
            COUNT(*),
            SUM(CASE WHEN status='closed_win' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status='closed_loss' THEN 1 ELSE 0 END),
            ROUND(SUM(COALESCE(pnl,0)), 4) AS total_pnl,
            ROUND(AVG(COALESCE(pnl,0)), 4),
            ROUND(AVG(CASE WHEN cost>0 THEN pnl/cost*100 END), 2),
            ROUND(AVG(entry_price), 4),
            ROUND(AVG(time_remaining_s), 1)
        FROM paper_trades
        WHERE status IN ('closed_win','closed_loss') {extra}
        GROUP BY strategy_type, exit_reason
        ORDER BY strategy_type, 6 DESC
    """, params)

    last_strat = None
    for ri, row in enumerate(rows, start=2):
        r = list(row)
        wins, losses = r[3] or 0, r[4] or 0
        out = [r[0], r[1], r[2], wins, losses, _win_rate(wins, losses),
               r[5], r[6], r[7], r[8], r[9]]
        ws.append(out)
        if r[0] != last_strat and last_strat is not None:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ri, column=col).fill = SUBHEAD_FILL
        last_strat = r[0]
        _reason_cell(ws.cell(row=ri, column=2))
        for col in (7, 8, 9):
            _pnl_cell(ws.cell(row=ri, column=col))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 5: Time Remaining Analysis
# ---------------------------------------------------------------------------

def _build_time_remaining_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Time At Exit Analysis")
    headers = [
        "Time Bucket", "Trades", "Wins", "Losses", "Win Rate %",
        "Total P&L $", "Avg P&L $", "Avg P&L %", "Avg Entry $",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    extra = "AND timestamp LIKE ?" if date_str else ""
    params = (f"{date_str}%",) if date_str else ()
    rows = conn.execute(f"""
        SELECT
            CASE
                WHEN time_remaining_s IS NULL THEN 'Market Resolution'
                WHEN time_remaining_s <= 0    THEN '0s (at close)'
                WHEN time_remaining_s <= 30   THEN '1-30s'
                WHEN time_remaining_s <= 60   THEN '31-60s'
                WHEN time_remaining_s <= 90   THEN '61-90s'
                WHEN time_remaining_s <= 120  THEN '91-120s'
                ELSE '120s+'
            END as bucket,
            COUNT(*),
            SUM(CASE WHEN status='closed_win' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status='closed_loss' THEN 1 ELSE 0 END),
            ROUND(SUM(COALESCE(pnl,0)), 4),
            ROUND(AVG(COALESCE(pnl,0)), 4),
            ROUND(AVG(CASE WHEN cost>0 THEN pnl/cost*100 END), 2),
            ROUND(AVG(entry_price), 4)
        FROM paper_trades
        WHERE status IN ('closed_win','closed_loss') {extra}
        GROUP BY bucket
    """, params)

    for ri, row in enumerate(rows, start=2):
        r = list(row)
        wins, losses = r[2] or 0, r[3] or 0
        out = [r[0], r[1], wins, losses, _win_rate(wins, losses),
               r[4], r[5], r[6], r[7]]
        ws.append(out)
        for col in (6, 7, 8):
            _pnl_cell(ws.cell(row=ri, column=col))
        _zebra(ws, ri, len(headers))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 6: Entry Price Level Analysis
# ---------------------------------------------------------------------------

def _build_price_level_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Entry Price Analysis")
    headers = [
        "Entry Price Bucket", "Trades", "Wins", "Losses", "Win Rate %",
        "Total P&L $", "Avg P&L $", "Avg P&L %", "Avg Exit $",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    extra = "AND timestamp LIKE ?" if date_str else ""
    params = (f"{date_str}%",) if date_str else ()
    rows = conn.execute(f"""
        SELECT
            CASE
                WHEN entry_price < 0.15 THEN 'Under 0.15'
                WHEN entry_price < 0.25 THEN '0.15 - 0.25'
                WHEN entry_price < 0.35 THEN '0.25 - 0.35'
                WHEN entry_price < 0.45 THEN '0.35 - 0.45'
                WHEN entry_price < 0.55 THEN '0.45 - 0.55'
                ELSE '0.55+'
            END as bucket,
            COUNT(*),
            SUM(CASE WHEN status='closed_win' THEN 1 ELSE 0 END),
            SUM(CASE WHEN status='closed_loss' THEN 1 ELSE 0 END),
            ROUND(SUM(COALESCE(pnl,0)), 4),
            ROUND(AVG(COALESCE(pnl,0)), 4),
            ROUND(AVG(CASE WHEN cost>0 THEN pnl/cost*100 END), 2),
            ROUND(AVG(CASE WHEN status IN ('closed_win','closed_loss') THEN exit_price END), 4)
        FROM paper_trades
        WHERE status IN ('closed_win','closed_loss') {extra}
        GROUP BY bucket ORDER BY MIN(entry_price)
    """, params)

    for ri, row in enumerate(rows, start=2):
        r = list(row)
        wins, losses = r[2] or 0, r[3] or 0
        out = [r[0], r[1], wins, losses, _win_rate(wins, losses),
               r[4], r[5], r[6], r[7]]
        ws.append(out)
        for col in (6, 7, 8):
            _pnl_cell(ws.cell(row=ri, column=col))
        _zebra(ws, ri, len(headers))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 7: Daily Summary
# ---------------------------------------------------------------------------

def _build_daily_summary_sheet(wb, conn):
    ws = wb.create_sheet("Daily Summary")
    headers = [
        "Date", "Evaluations", "Signals", "Skips",
        "Paper Trades", "Paper Wins", "Paper Losses", "Paper Win Rate %", "Paper P&L $", "Open Positions",
        "Live Trades", "Live Successful", "Live Aborted", "Live P&L $",
        "Missed Profit $ (Skipped Winners)",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    dates = conn.execute("""
        SELECT DATE(timestamp) FROM decisions GROUP BY DATE(timestamp) ORDER BY DATE(timestamp) DESC
    """).fetchall()

    for ri, (date,) in enumerate(dates, start=2):
        dec = conn.execute("""
            SELECT COUNT(*),
                   SUM(CASE WHEN action IN ('TRADE','DRY_RUN_SIGNAL') THEN 1 ELSE 0 END),
                   SUM(CASE WHEN action='SKIP' THEN 1 ELSE 0 END)
            FROM decisions WHERE DATE(timestamp) = ?
        """, (date,)).fetchone()

        paper = conn.execute("""
            SELECT COUNT(*),
                   SUM(CASE WHEN status='closed_win' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN status='closed_loss' THEN 1 ELSE 0 END),
                   ROUND(SUM(COALESCE(pnl,0)), 4),
                   SUM(CASE WHEN status='open' THEN 1 ELSE 0 END)
            FROM paper_trades WHERE DATE(timestamp) = ?
        """, (date,)).fetchone()

        live = conn.execute("""
            SELECT COUNT(*),
                   SUM(CASE WHEN status='success' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN status IN ('aborted','partial_abort') THEN 1 ELSE 0 END),
                   ROUND(COALESCE(SUM(do.actual_profit), 0), 4)
            FROM trades t
            LEFT JOIN decision_outcomes do ON do.decision_id = t.decision_id
            WHERE DATE(t.timestamp) = ?
        """, (date,)).fetchone()

        missed = conn.execute("""
            SELECT ROUND(COALESCE(SUM(do.hypothetical_profit), 0), 4)
            FROM decisions d
            JOIN decision_outcomes do ON do.decision_id = d.id
            WHERE DATE(d.timestamp) = ? AND d.action='SKIP' AND do.would_have_profited=1
        """, (date,)).fetchone()

        p_total, p_wins, p_losses, p_pnl, p_open = paper or (0, 0, 0, 0, 0)
        l_total, l_win, l_abort, l_pnl = live or (0, 0, 0, 0)

        row = [
            date, dec[0], dec[1], dec[2],
            p_total or 0, p_wins or 0, p_losses or 0,
            _win_rate(p_wins, p_losses), round(p_pnl or 0, 4), p_open or 0,
            l_total or 0, l_win or 0, l_abort or 0, round(l_pnl or 0, 4),
            missed[0] if missed else 0,
        ]
        ws.append(row)
        _pnl_cell(ws.cell(row=ri, column=9))    # Paper P&L
        _pnl_cell(ws.cell(row=ri, column=14))   # Live P&L
        _zebra(ws, ri, len(headers))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 8: Live Trades
# ---------------------------------------------------------------------------

def _build_trades_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Live Trades")
    headers = [
        "Timestamp", "Coin", "Strategy", "Side", "Market",
        "UP Price $", "DOWN Price $", "Entry $", "Target $",
        "Combined Cost", "Fees $", "Shares", "Status", "Abort Cost $", "Actual P&L $",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    where = "WHERE t.timestamp LIKE ?" if date_str else ""
    params = (f"{date_str}%",) if date_str else ()
    for ri, row in enumerate(conn.execute(f"""
        SELECT t.timestamp, t.coin, t.strategy_type, t.side, t.market_slug,
               t.up_price, t.down_price, t.entry_price, t.target_price,
               t.combined_cost, t.fee_total, t.shares, t.status, t.abort_cost,
               do.actual_profit
        FROM trades t
        LEFT JOIN decision_outcomes do ON do.decision_id = t.decision_id
        {where} ORDER BY t.id DESC
    """, params), start=2):
        r = list(row)
        ws.append(r)
        status = str(r[12] or "").lower()
        row_fill = (WIN_FILL if status == "success" else
                    LOSS_FILL if status in ("aborted", "partial_abort", "failed") else None)
        if row_fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ri, column=col).fill = row_fill
        _pnl_cell(ws.cell(row=ri, column=15))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 9: All Decisions
# ---------------------------------------------------------------------------

def _build_decisions_sheet(wb, conn, date_str):
    ws = wb.create_sheet("All Decisions")
    headers = [
        "Timestamp", "Coin", "Strategy", "Market",
        "UP Ask", "DOWN Ask", "Combined",
        "Coin Price $", "Velocity 30s %", "Velocity 60s %",
        "UP Depth $", "DOWN Depth $", "Time Left (s)",
        "Feeds", "Passed Filters", "Failed Filters",
        "Action", "Reason", "Would Have Profited", "Hypothetical P&L $",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    where = "WHERE d.timestamp LIKE ?" if date_str else ""
    params = (f"{date_str}%",) if date_str else ()
    for ri, row in enumerate(conn.execute(f"""
        SELECT d.timestamp, d.coin, d.strategy_type, d.market_slug,
               d.up_best_ask, d.down_best_ask, d.combined_cost,
               d.btc_price, d.coin_velocity_30s, d.coin_velocity_60s,
               d.up_depth_usd, d.down_depth_usd, d.time_remaining_s,
               d.feed_count, d.filter_passed, d.filter_failed,
               d.action, d.reason,
               do.would_have_profited, do.hypothetical_profit
        FROM decisions d
        LEFT JOIN decision_outcomes do ON do.decision_id = d.id
        {where} ORDER BY d.id DESC LIMIT 10000
    """, params), start=2):
        r = list(row)
        ws.append(r)
        action = str(r[16] or "").upper()
        if action in ("TRADE", "DRY_RUN_SIGNAL"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=ri, column=col).fill = WIN_FILL
        _pnl_cell(ws.cell(row=ri, column=20))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 10: Signals
# ---------------------------------------------------------------------------

def _build_signals_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Signals")
    headers = [
        "Timestamp", "Run ID", "Coin", "Strategy", "Mode", "Action", "Suppressed Reason",
        "Reason", "Signal Context", "Overlap Count",
        "Entry $", "Target $", "Expected Profit/Share $",
        "Signal Score", "Score Velocity", "Score Entry", "Score Depth", "Score Spread", "Score Time", "Score Balance",
        "Time Left (s)", "Velocity 30s %", "Velocity 60s %",
        "Entry Bid $", "Entry Ask $", "Entry Spread $",
        "Entry Depth $", "Opposite Depth $", "Depth Ratio",
        "Resignal Cooldown (s)", "Min Price Improvement", "Last Signal Age (s)",
        "Passed Filters", "Failed Filters",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    where = "WHERE timestamp LIKE ? AND action IN ('DRY_RUN_SIGNAL','TRADE','SUPPRESSED')" if date_str else "WHERE action IN ('DRY_RUN_SIGNAL','TRADE','SUPPRESSED')"
    params = (f"{date_str}%",) if date_str else ()
    rows = conn.execute(f"""
        SELECT timestamp, run_id, coin, strategy_type, mode, action, suppressed_reason,
               reason, signal_context, signal_overlap_count,
               entry_price, target_price, expected_profit_per_share,
               signal_score, score_velocity, score_entry, score_depth, score_spread, score_time, score_balance,
               time_remaining_s, coin_velocity_30s, coin_velocity_60s,
               entry_bid, entry_ask, entry_spread,
               entry_depth_side_usd, opposite_depth_usd, depth_ratio,
               resignal_cooldown_s, min_price_improvement, last_signal_age_s,
               filter_passed, filter_failed
        FROM decisions
        {where}
        ORDER BY id DESC LIMIT 10000
    """, params)

    for ri, row in enumerate(rows, start=2):
        ws.append(list(row))
        _action_cell(ws.cell(row=ri, column=6))
        suppressed_cell = ws.cell(row=ri, column=7)
        if suppressed_cell.value:
            suppressed_cell.fill = LOSS_FILL
            suppressed_cell.font = LOSS_FONT
        _pnl_cell(ws.cell(row=ri, column=13))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 11: Filter Performance
# ---------------------------------------------------------------------------

def _build_filter_performance_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Filter Performance")
    headers = [
        "Filter", "Times Passed", "Times Failed", "Reject Rate %",
        "Correct Rejections", "Missed Winners", "Accuracy %",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    where = "WHERE timestamp LIKE ?" if date_str else ""
    params = (f"{date_str}%",) if date_str else ()
    rows = conn.execute(
        f"SELECT filter_passed, filter_failed FROM decisions {where}", params
    ).fetchall()

    counts: dict[str, dict] = {}
    for passed_str, failed_str in rows:
        for name in (passed_str or "").split(","):
            name = name.strip()
            if name:
                counts.setdefault(name, {"passed": 0, "failed": 0})
                counts[name]["passed"] += 1
        for name in (failed_str or "").split(","):
            name = name.strip()
            if name:
                counts.setdefault(name, {"passed": 0, "failed": 0})
                counts[name]["failed"] += 1

    for fname in counts:
        extra = "AND d.timestamp LIKE ?" if date_str else ""
        acc_params = [f"%{fname}%"] + ([f"{date_str}%"] if date_str else [])
        result = conn.execute(f"""
            SELECT
                SUM(CASE WHEN do.would_have_profited=0 THEN 1 ELSE 0 END),
                SUM(CASE WHEN do.would_have_profited=1 THEN 1 ELSE 0 END)
            FROM decisions d
            JOIN decision_outcomes do ON do.decision_id=d.id
            WHERE d.filter_failed LIKE ? {extra}
        """, acc_params).fetchone()
        counts[fname]["correct"] = result[0] or 0 if result else 0
        counts[fname]["missed"]  = result[1] or 0 if result else 0

    for ri, (name, c) in enumerate(sorted(counts.items()), start=2):
        total = c["passed"] + c["failed"]
        reject_pct = round(c["failed"] / total * 100, 1) if total else 0
        correct, missed = c.get("correct", 0), c.get("missed", 0)
        acc_total = correct + missed
        accuracy = round(correct / acc_total * 100, 1) if acc_total else 0
        ws.append([name, c["passed"], c["failed"], reject_pct, correct, missed, accuracy])
        acc_cell = ws.cell(row=ri, column=7)
        if accuracy >= 70:
            acc_cell.fill = WIN_FILL
            acc_cell.font = WIN_FONT
        elif accuracy < 50 and acc_total > 0:
            acc_cell.fill = LOSS_FILL
            acc_cell.font = LOSS_FONT
        _zebra(ws, ri, len(headers))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 12: Skipped Winners
# ---------------------------------------------------------------------------

def _build_skipped_winners_sheet(wb, conn, date_str):
    ws = wb.create_sheet("Skipped Winners")
    headers = [
        "Timestamp", "Coin", "Market", "UP Ask", "DOWN Ask", "Combined $",
        "Failed Filters", "Reason", "Hypothetical P&L $", "Winner",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    _freeze(ws)

    params = []
    date_clause = ""
    if date_str:
        date_clause = "AND d.timestamp LIKE ?"
        params.append(f"{date_str}%")

    for ri, row in enumerate(conn.execute(f"""
        SELECT d.timestamp, d.coin, d.market_slug,
               d.up_best_ask, d.down_best_ask, d.combined_cost,
               d.filter_failed, d.reason,
               do.hypothetical_profit, o.winner
        FROM decisions d
        JOIN decision_outcomes do ON do.decision_id=d.id
        JOIN outcomes o ON do.outcome_id=o.id
        WHERE d.action='SKIP' AND do.would_have_profited=1 {date_clause}
        ORDER BY do.hypothetical_profit DESC LIMIT 5000
    """, params), start=2):
        ws.append(list(row))
        for col in range(1, len(headers) + 1):
            ws.cell(row=ri, column=col).fill = LOSS_FILL   # missed = shown red
        _pnl_cell(ws.cell(row=ri, column=9))

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
