import csv
import gzip
import sys
import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "edec_bot"))

from bot.archive import _build_session_excel_export, _session_export_folder_name


class SessionExcelExportTests(unittest.TestCase):
    def test_session_export_folder_name_is_stable(self):
        folder = _session_export_folder_name(
            datetime(2026, 4, 19, 16, 55, 0, tzinfo=timezone.utc),
            "EDEC-BOT",
        )

        self.assertEqual(folder, "2026-04-19_165500_EDEC-BOT_session_export")

    def test_build_session_excel_export_creates_summary_and_data_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            trades_path = output_dir / "session_trades.csv.gz"
            signals_path = output_dir / "session_signals.csv.gz"

            with gzip.open(trades_path, "wt", encoding="utf-8", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(["trade_id", "coin", "pnl"])
                writer.writerow([1, "btc", "12.34"])

            with gzip.open(signals_path, "wt", encoding="utf-8", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(["decision_id", "coin", "action"])
                writer.writerow([9, "btc", "TRADE"])

            excel_path = _build_session_excel_export(
                output_dir=str(output_dir),
                label="EDEC-BOT",
                now_utc=datetime(2026, 4, 19, 16, 55, 0, tzinfo=timezone.utc),
                reset_at="2026-04-19T15:00:00",
                trades_path=str(trades_path),
                signals_path=str(signals_path),
                trade_count=1,
                signal_count=1,
            )

            self.assertTrue(Path(excel_path).exists())
            wb = load_workbook(excel_path, read_only=True)
            self.assertEqual(wb.sheetnames, ["Summary", "Session Trades", "Session Signals"])
            summary = wb["Summary"]
            self.assertEqual(summary["A2"].value, "Label")
            self.assertEqual(summary["B2"].value, "EDEC-BOT")
            self.assertEqual(wb["Session Trades"]["A2"].value, "1")
            self.assertEqual(wb["Session Signals"]["C2"].value, "TRADE")
            wb.close()


if __name__ == "__main__":
    unittest.main()
